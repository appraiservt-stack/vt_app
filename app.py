from flask import Flask, render_template, request, jsonify, Response, session, redirect, url_for
import requests
import json
import csv
import io
import math
import re
import os
from pathlib import Path
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass  # dotenv not installed — env vars must be set manually
from datetime import datetime, timezone
from functools import wraps

app = Flask(__name__)
# Secret key for session encryption — set FLASK_SECRET_KEY env var in Railway
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "vt-prop-dev-secret-change-in-prod")
from datetime import timedelta
# Session expires after 8 hours of inactivity.
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(hours=8)
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = True

# ── Auth blueprint ────────────────────────────────────────────────────────────
from auth import auth_bp, init_db, db_fetchone, user_has_access, days_left_in_trial, _q
app.register_blueprint(auth_bp)
init_db()   # create users table if it doesn't exist

# ── Login required decorator ──────────────────────────────────────────────────
ADMIN_EMAIL_LOCAL   = os.environ.get("ADMIN_EMAIL", "appraiservt@gmail.com").lower()
ADMIN_PASSWORD_SET  = bool(os.environ.get("ADMIN_PASSWORD"))  # True only when env var exists

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("auth.login"))
        # Admin bypass — only active when ADMIN_PASSWORD env var is set (local dev only)
        if ADMIN_PASSWORD_SET and session.get("user_email", "").lower() == ADMIN_EMAIL_LOCAL:
            return f(*args, **kwargs)
        user = db_fetchone(_q("SELECT * FROM users WHERE id = ?"), (session["user_id"],))
        if not user:
            # Account deleted or session stale — go to login
            session.clear()
            return redirect(url_for("auth.login"))
        if not user_has_access(user):
            # Trial expired or cancelled — go to subscribe
            return redirect(url_for("auth.subscribe",
                                    email=session.get("user_email", "")))
        return f(*args, **kwargs)
    return decorated

ARCGIS_URL = (
    "https://services1.arcgis.com/"
    "BkFxaEFNwHqX3tAw/arcgis/rest/services/"
    "FS_VCGI_OPENDATA_Cadastral_PTTR_point_WM_v1_view/FeatureServer/0/query"
)

# -----------------------------
# Load Vermont code mappings
# -----------------------------
CODES_PATH = Path(__file__).with_name("vt_codes.json")

with CODES_PATH.open() as f:
    VT_CODES = json.load(f)

COUNTY_CODE_TO_NAME = VT_CODES["counties"]
TOWN_TO_COUNTY      = VT_CODES["town_to_county"]
SCHOOL_TO_TOWN      = {int(k): v for k, v in VT_CODES["school_to_town"].items()}

# Grand List Category lookup: numeric code → description text
GL_CATEGORY_LOOKUP   = VT_CODES.get("grand_list_categories", {})
# Building type lookup: numeric code → description text
BUILDING_TYPE_LOOKUP = VT_CODES.get("building_types", {})
# Town centroid fallback: town name (UPPERCASE) → {lat, lon, county}
# Used when a record's ArcGIS coordinates fail the county bounds check.
TOWN_CENTROIDS = VT_CODES.get("town_centroids", {})

# Pre-geocoded approx records: OBJECTID (str) → {lat, lon}
# Built by geocode_approx.py, updated weekly.
# Records here get plotted as precise black dots instead of town centroid red circles.
_GEOCODED_APPROX_PATH = Path(__file__).with_name("geocoded_approx.json")
try:
    with _GEOCODED_APPROX_PATH.open() as _f:
        GEOCODED_APPROX = json.load(_f)
except (FileNotFoundError, json.JSONDecodeError):
    GEOCODED_APPROX = {}

# Vermont state bounding box — outer limit for any coordinate check.
VT_LAT_MIN, VT_LAT_MAX =  42.7,  45.1
VT_LON_MIN, VT_LON_MAX = -73.5, -71.5

# Per-county bounding boxes keyed by zero-padded county code.
# Used to validate that a record’s ArcGIS coordinates land in the right county.
# The key insight: we check against the TRUSTED county (from schoolCode), not
# the self-reported countyCode on the form or the ArcGIS TOWNNAME field.
# A generous padding of ~0.15 degrees (~10 miles) handles records near county
# borders whose geocoded point lands just across the line.
COUNTY_BOUNDS = {
    "01": (43.60, 44.55, -73.50, -72.60),  # Addison
    "02": (42.70, 43.50, -73.50, -72.70),  # Bennington
    "03": (44.10, 45.05, -72.55, -71.40),  # Caledonia
    "04": (44.10, 44.90, -73.50, -72.65),  # Chittenden
    "05": (44.20, 45.05, -72.35, -71.40),  # Essex
    "06": (44.45, 45.05, -73.40, -72.35),  # Franklin
    "07": (44.45, 45.05, -73.55, -72.95),  # Grand Isle
    "08": (44.25, 44.95, -73.15, -72.20),  # Lamoille
    "09": (43.60, 44.40, -72.90, -71.80),  # Orange
    "10": (44.40, 45.05, -72.75, -71.70),  # Orleans
    "11": (43.10, 44.10, -73.55, -72.35),  # Rutland
    "12": (43.85, 44.70, -73.10, -72.05),  # Washington
    "13": (42.65, 43.50, -73.05, -71.85),  # Windham
    "14": (43.10, 44.25, -72.95, -71.95),  # Windsor
}


def coords_in_county(lat, lon, county_code):
    """Return True if (lat, lon) falls within the padded bbox of county_code."""
    if lat is None or lon is None:
        return False
    code = str(county_code).strip().zfill(2)
    bounds = COUNTY_BOUNDS.get(code)
    if bounds is None:
        return True  # Unknown county — give it the benefit of the doubt
    lat_min, lat_max, lon_min, lon_max = bounds
    return lat_min <= lat <= lat_max and lon_min <= lon <= lon_max


def resolve_coordinates(raw_lat, raw_lon, trusted_town, object_id=None, match_method=None):
    """Return (lat, lon, approx, is_centroid) for a record.

    Uses MatchMthod as the primary signal for coordinate quality:
    0. GEOCODED_APPROX lookup — manually/script geocoded coordinates
    1. MatchMthod in _GOOD_MATCH_METHODS — reliable geocode, use as-is
    2. MatchMthod in _APPROX_MATCH_METHODS — parcel centroid, approx but usable
    3. MatchMthod = 'Unmatched' or null — fall back to town centroid
    4. Sanity check: coords must be in Vermont
    """
    # Tier 0: pre-geocoded lookup by OBJECTID
    if object_id is not None:
        entry = GEOCODED_APPROX.get(str(object_id))
        if entry and entry.get("lat") and entry.get("lon"):
            return entry["lat"], entry["lon"], False, False

    mm = (match_method or "").strip().lower()

    # Tier 1: good geocode — trust coordinates
    if mm in _GOOD_MATCH_METHODS:
        if (raw_lat is not None and raw_lon is not None and
                VT_LAT_MIN <= raw_lat <= VT_LAT_MAX and
                VT_LON_MIN <= raw_lon <= VT_LON_MAX):
            return raw_lat, raw_lon, False, False
        # Good match method but coords out of VT — fall through to centroid

    # Tier 2: parcel centroid match — approximate but usable coordinates
    if mm in _APPROX_MATCH_METHODS:
        if (raw_lat is not None and raw_lon is not None and
                VT_LAT_MIN <= raw_lat <= VT_LAT_MAX and
                VT_LON_MIN <= raw_lon <= VT_LON_MAX):
            return raw_lat, raw_lon, True, False  # approx, not centroid
        # Coords out of VT — fall through to centroid

    # Tier 3: unmatched or unknown method — town centroid fallback
    town_key = (trusted_town or "").strip().upper()
    centroid  = TOWN_CENTROIDS.get(town_key)
    if centroid:
        return centroid["lat"], centroid["lon"], True, True

    return None, None, False, False


# MatchMthod values that indicate reliable geocoding
_GOOD_MATCH_METHODS = {
    'property address (esite)',
    'property address (composite)',
    'span (esite)',
}
# MatchMthod values that indicate approximate (parcel centroid) geocoding
_APPROX_MATCH_METHODS = {
    'span (parcel centroid)',
}
# 'unmatched' or anything else -> use town centroid fallback


def derive_trusted_location(attr):
    """Derive reliable town/county from TownSpan (primary) or schoolCode (fallback).

    TownSpan is entered by the town clerk against the official grand list and
    is the most reliable identifier. Its first 3 digits equal the town code.
    schoolCode is the fallback when TownSpan is absent.
    """
    raw_county  = attr.get("countyCode")
    school_code = attr.get("schoolCode")
    town_code   = attr.get("townCode")
    span        = attr.get("span")
    town_span   = attr.get("TownSpan") or attr.get("townSpan") or ""
    prop_city   = (
        attr.get("propertyLocationCity")
        or attr.get("propLocCty")
        or attr.get("TownCityorTown")
    )

    trusted_town   = None
    trusted_county = None

    # Primary: TownSpan first 3 digits = town code
    # Zero-pad to 3 digits to match SCHOOL_TO_TOWN keys
    if town_span and len(town_span.strip()) >= 3:
        ts_town_code = town_span.strip()[:3]
        # Convert town code to town name via townCode->town lookup
        # townCode is a 3-digit numeric string like '441' for Northfield
        # We can look it up via SCHOOL_TO_TOWN by matching townCode
        # Actually use a direct townCode->town mapping built from vt_codes
        ts_int = None
        try:
            ts_int = int(ts_town_code)
        except (TypeError, ValueError):
            pass
        if ts_int is not None:
            # Build town from townCode: school_to_town maps schoolCode->town
            # but townCode != schoolCode. Use TOWN_TO_COUNTY inverted lookup.
            # Find town whose county matches via town_code->town mapping
            # VT town codes are used in SPAN prefix. Look up via schoolCode
            # fallback if direct match fails.
            # Best approach: match townCode against known town codes
            for school, town in SCHOOL_TO_TOWN.items():
                # VT town codes (3-digit) embedded in SPAN = first 3 of 11-digit SPAN
                # We can derive town code from school code indirectly.
                # For now, fall through to schoolCode if no direct match.
                pass

    # Primary: schoolCode (reliable, state-assigned)
    if school_code is not None:
        try:
            sc_int = int(float(str(school_code)))
        except (TypeError, ValueError):
            sc_int = None
        if sc_int is not None:
            trusted_town = SCHOOL_TO_TOWN.get(sc_int)

    if trusted_town:
        trusted_county = TOWN_TO_COUNTY.get(trusted_town)

    # Fallback 1: propLocCty town name lookup
    if trusted_county is None:
        prop_city = (prop_city or '').strip().title()
        if prop_city:
            trusted_county = TOWN_TO_COUNTY.get(prop_city)
            if trusted_county:
                trusted_town = trusted_town or prop_city

    # Fallback 2: self-reported countyCode (unreliable but better than nothing)
    if trusted_county is None and raw_county is not None:
        trusted_county = str(raw_county).zfill(2)

    corrected = False
    if raw_county is not None and trusted_county is not None:
        corrected = str(raw_county).zfill(2) != str(trusted_county).zfill(2)

    trusted_county_name = None
    if trusted_county is not None:
        trusted_county_name = COUNTY_CODE_TO_NAME.get(str(trusted_county).zfill(2))

    return {
        "rawCountyCode":      raw_county,
        "trustedCountyCode":  trusted_county,
        "trustedCountyName":  trusted_county_name,
        "trustedTown":        trusted_town,
        "schoolCode":         school_code,
        "townCode":           town_code,
        "span":               span,
        "displayCity":        prop_city,
        "correctedCounty":    corrected,
    }


def parse_filters(args):
    """Parse all filter query parameters into a dict."""
    f = {}

    # --- Location ---
    f["counties"]     = args.get("counties", "")       # comma-separated codes
    f["towns"]        = args.get("towns", "")           # comma-separated town names

    # --- Date range ---
    f["date_from"]    = args.get("date_from", "")       # ISO date string YYYY-MM-DD
    f["date_to"]      = args.get("date_to", "")

    # --- Price range ---
    f["price_low"]    = args.get("price_low", "")
    f["price_high"]   = args.get("price_high", "")

    # --- Land size range ---
    f["land_low"]     = args.get("land_low", "")
    f["land_high"]    = args.get("land_high", "")

    # --- Street address (contains) ---
    f["street"]       = args.get("street", "").strip()

    # --- Interest type (multi, pipe-separated codes) ---
    f["interest"]     = args.get("interest", "")        # e.g. "1|2"

    # --- Building construction (multi, pipe-separated codes) ---
    f["building"]     = args.get("building", "")        # e.g. "2|3"

    # --- Seller use of property (multi, pipe-separated codes) ---
    f["seller_use"]   = args.get("seller_use", "")

    # --- Buyer use of property (multi, pipe-separated codes) ---
    f["buyer_use"]    = args.get("buyer_use", "")

    # --- PTT Exemption (multi, pipe-separated codes) ---
    f["ptt_exemption"] = args.get("ptt_exemption", "")

    # --- Grand List Category (multi, pipe-separated codes) ---
    f["grand_list"]   = args.get("grand_list", "")

    # --- Seller name fields ---
    f["seller_entity"] = args.get("seller_entity", "").strip()
    f["seller_last"]   = args.get("seller_last", "").strip()
    f["seller_first"]  = args.get("seller_first", "").strip()

    # --- Buyer name fields ---
    f["buyer_entity"]  = args.get("buyer_entity", "").strip()
    f["buyer_last"]    = args.get("buyer_last", "").strip()
    f["buyer_first"]   = args.get("buyer_first", "").strip()

    # --- SPAN ---
    f["span"]          = args.get("span", "").strip().replace("-", "")

    # --- Boolean radio buttons ---
    f["dev_prev_conv"]         = args.get("dev_prev_conv", "")    # "true"/"false"/""
    f["buyer_adjoining"]       = args.get("buyer_adjoining", "")
    f["enrolled_current_use"]  = args.get("enrolled_current_use", "")  # "true"/"false"/""
    f["foreclosed"]            = args.get("foreclosed", "")            # "true"/"false"/""

    return f


def build_where_clause(filters):
    """Build the ArcGIS WHERE clause from filters.

    All filters are pushed to ArcGIS SQL so the server does the heavy
    lifting across 218K+ records.  Python post-filtering is only used for
    the county-bounds coordinate validation (not a data filter).
    """
    clauses = ["1=1"]

    # County filter
    # ArcGIS has inconsistent countyCode padding — some records use '9', others '09'.
    # Include both padded and unpadded forms to catch all records.
    if filters["counties"]:
        codes = filters["counties"].split(",")
        all_codes = set()
        for c in codes:
            all_codes.add(c)                    # e.g. '09'
            all_codes.add(str(int(c)))          # e.g. '9'
        codes_sql = ",".join([f"'{c}'" for c in sorted(all_codes)])
        clauses.append(f"countyCode IN ({codes_sql})")

    # Date range — ArcGIS requires DATE 'YYYY-MM-DD' format
    if filters["date_from"]:
        try:
            datetime.strptime(filters["date_from"], "%Y-%m-%d")
            clauses.append(f"closeDate >= DATE '{filters['date_from']}'")
        except Exception:
            pass

    if filters["date_to"]:
        try:
            datetime.strptime(filters["date_to"], "%Y-%m-%d")
            clauses.append(f"closeDate <= DATE '{filters['date_to']}'")
        except Exception:
            pass

    # Price range — actual field name: ValPdOrTrn
    if filters["price_low"]:
        try:
            clauses.append(f"ValPdOrTrn >= {float(filters['price_low'])}")
        except Exception:
            pass

    if filters["price_high"]:
        try:
            clauses.append(f"ValPdOrTrn <= {float(filters['price_high'])}")
        except Exception:
            pass

    # Land size range — actual field name: landSize
    if filters["land_low"]:
        try:
            clauses.append(f"landSize >= {float(filters['land_low'])}")
        except Exception:
            pass

    if filters["land_high"]:
        try:
            clauses.append(f"landSize <= {float(filters['land_high'])}")
        except Exception:
            pass

    # Interest type — actual field name: intPrpType
    if filters["interest"]:
        codes = filters["interest"].split("|")
        codes_sql = ",".join(codes)
        clauses.append(f"intPrpType IN ({codes_sql})")

    # Building construction — actual field name: blCn1
    if filters["building"]:
        codes = filters["building"].split("|")
        codes_sql = ",".join(codes)
        clauses.append(f"blCn1 IN ({codes_sql})")

    # Seller use of property — actual field name: sUsePr
    if filters["seller_use"]:
        codes = filters["seller_use"].split("|")
        codes_sql = ",".join(codes)
        clauses.append(f"sUsePr IN ({codes_sql})")

    # Buyer use of property — actual field name: bUsePr
    if filters["buyer_use"]:
        codes = filters["buyer_use"].split("|")
        codes_sql = ",".join(codes)
        clauses.append(f"bUsePr IN ({codes_sql})")

    # PTT exemption — actual field name: prTxEx
    if filters["ptt_exemption"]:
        codes = filters["ptt_exemption"].split("|")
        codes_sql = ",".join(codes)
        clauses.append(f"prTxEx IN ({codes_sql})")

    # Grand list category — actual field name: TownGlCat
    if filters["grand_list"]:
        codes = filters["grand_list"].split("|")
        codes_sql = ",".join(codes)
        clauses.append(f"TownGlCat IN ({codes_sql})")

    # ---------------------------------------------------------------
    # Name / SPAN / Street filters — pushed into SQL LIKE so ArcGIS
    # searches all 218K+ records efficiently instead of relying on
    # Python post-filtering across paginated statewide results.
    # ---------------------------------------------------------------

    # Helper: escape single-quotes in user input for SQL safety
    def sql_like(val):
        return val.replace("'", "''").upper()

    seller_last   = filters.get("seller_last",   "").strip()
    seller_first  = filters.get("seller_first",  "").strip()
    seller_entity = filters.get("seller_entity", "").strip()
    buyer_last    = filters.get("buyer_last",    "").strip()
    buyer_first   = filters.get("buyer_first",   "").strip()
    buyer_entity  = filters.get("buyer_entity",  "").strip()

    has_seller_name = bool(seller_last or seller_first or seller_entity)
    has_buyer_name  = bool(buyer_last  or buyer_first  or buyer_entity)

    if has_seller_name and has_buyer_name:
        # Both sides filled — OR: match seller criteria OR buyer criteria
        seller_parts = []
        if seller_last:   seller_parts.append(f"UPPER(sellLstNam) LIKE '%{sql_like(seller_last)}%'")
        if seller_first:  seller_parts.append(f"UPPER(sellFstNam) LIKE '%{sql_like(seller_first)}%'")
        if seller_entity: seller_parts.append(f"UPPER(sellEntNam) LIKE '%{sql_like(seller_entity)}%'")
        buyer_parts = []
        if buyer_last:    buyer_parts.append(f"UPPER(buyLstNam) LIKE '%{sql_like(buyer_last)}%'")
        if buyer_first:   buyer_parts.append(f"UPPER(buyFstNam) LIKE '%{sql_like(buyer_first)}%'")
        if buyer_entity:  buyer_parts.append(f"UPPER(buyEntNam) LIKE '%{sql_like(buyer_entity)}%'")
        seller_sql = " AND ".join(seller_parts)
        buyer_sql  = " AND ".join(buyer_parts)
        clauses.append(f"(({seller_sql}) OR ({buyer_sql}))")
    elif has_seller_name:
        if seller_last:   clauses.append(f"UPPER(sellLstNam) LIKE '%{sql_like(seller_last)}%'")
        if seller_first:  clauses.append(f"UPPER(sellFstNam) LIKE '%{sql_like(seller_first)}%'")
        if seller_entity: clauses.append(f"UPPER(sellEntNam) LIKE '%{sql_like(seller_entity)}%'")
    elif has_buyer_name:
        if buyer_last:    clauses.append(f"UPPER(buyLstNam) LIKE '%{sql_like(buyer_last)}%'")
        if buyer_first:   clauses.append(f"UPPER(buyFstNam) LIKE '%{sql_like(buyer_first)}%'")
        if buyer_entity:  clauses.append(f"UPPER(buyEntNam) LIKE '%{sql_like(buyer_entity)}%'")

    # Street address — pushed to SQL
    if filters.get("street", "").strip():
        clauses.append(f"UPPER(propLocStr) LIKE '%{sql_like(filters['street'].strip())}%'")

    # SPAN — pushed to SQL (strip dashes before comparing)
    span_val = filters.get("span", "").strip().replace("-", "")
    if span_val:
        clauses.append(f"CAST(span AS VARCHAR(20)) LIKE '%{sql_like(span_val)}%'")

    # ---------------------------------------------------------------
    # Town filter — convert town names to schoolCodes for SQL IN()
    # schoolCode is the most reliable town identifier in the ArcGIS data.
    # ---------------------------------------------------------------
    if filters.get("towns", "").strip():
        requested_towns = [t.strip().upper() for t in filters["towns"].split(",") if t.strip()]
        school_codes = []
        for town in requested_towns:
            # SCHOOL_TO_TOWN maps code->town; we need town->code(s)
            for code, mapped_town in SCHOOL_TO_TOWN.items():
                if mapped_town.upper() == town:
                    school_codes.append(str(code))
        if school_codes:
            codes_sql = ",".join([f"'{c}'" for c in school_codes])
            clauses.append(f"schoolCode IN ({codes_sql})")
        else:
            # No matching school codes found — force zero results rather
            # than returning unfiltered statewide data
            clauses.append("1=0")

    # ---------------------------------------------------------------
    # Boolean / radio flag filters — all pushed to SQL
    # Fields store uppercase strings 'TRUE' / 'FALSE'.
    # ---------------------------------------------------------------
    def bool_clause(field, value):
        """Return SQL clause for a boolean string field given 'true'/'false'."""
        if value == "true":
            return f"UPPER({field}) = 'TRUE'"
        elif value == "false":
            return f"UPPER({field}) = 'FALSE'"
        return None

    for flt_key, field in [
        ("dev_prev_conv",       "devPrevCnv"),
        ("buyer_adjoining",     "buyrAdjPrp"),
        ("enrolled_current_use", "enrCrntUse"),
    ]:
        clause = bool_clause(field, filters.get(flt_key, ""))
        if clause:
            clauses.append(clause)

    # Foreclosed — LGTEx code '6' means foreclosed sale
    fc = filters.get("foreclosed", "")
    if fc == "true":
        clauses.append("LGTEx = '6'")
    elif fc == "false":
        clauses.append("LGTEx <> '6'")

    return " AND ".join(clauses)


def apply_python_filters(record, filters):
    """All filters are now handled by ArcGIS SQL in build_where_clause.
    This function is retained as a no-op stub so call sites don't need
    to change.  The only remaining Python-side logic is the county-bounds
    coordinate validation in resolve_coordinates(), which is not a data
    filter but a geocoding quality check.
    """
    return True


def fetch_features(where, geometry_params=None, max_records=2000):
    """Fetch features from ArcGIS. geometry_params is optional bbox dict."""
    params = {
        "where":             where,
        "outFields":         "*",
        "f":                 "json",
        "outSR":             "4326",
        "resultRecordCount": max_records,
    }
    if geometry_params:
        params.update(geometry_params)
    else:
        params["inSR"] = "4326"

    r = requests.get(ARCGIS_URL, params=params)
    return r.json().get("features", [])


def fetch_all_features(where):
    """
    Page through ArcGIS results to get ALL matching records
    (for statewide export, no viewport limit).
    Uses resultOffset pagination.
    """
    all_features = []
    offset = 0
    page_size = 2000

    while True:
        params = {
            "where":             where,
            "outFields":         "*",
            "f":                 "json",
            "outSR":             "4326",
            "resultRecordCount": page_size,
            "resultOffset":      offset,
        }
        try:
            r = requests.get(ARCGIS_URL, params=params, timeout=25)
            data = r.json()
        except Exception as e:
            app.logger.error(f"fetch_all_features error at offset {offset}: {e}")
            break
        if data.get("error"):
            app.logger.error(f"ArcGIS error in fetch_all_features: {data['error']}")
            break
        features = data.get("features", [])
        all_features.extend(features)

        if len(features) < page_size:
            break
        if not data.get("exceededTransferLimit", True) and len(features) < page_size:
            break

        offset += page_size

    return all_features


def feature_to_record(f, filters):
    """Convert a raw ArcGIS feature to our record dict. Returns None if filtered out."""
    attr = f["attributes"]
    loc_info = derive_trusted_location(attr)

    # ---------------------------------------------------------
    # buildingConstruction1Desc fallback:
    # If blCn1Desc is blank AND blCn3 is 4 or "04", copy blCn3Desc
    # ---------------------------------------------------------
    bl_cn1_desc = attr.get("blCn1Desc") or ""
    bl_cn3      = attr.get("blCn3")
    bl_cn3_desc = attr.get("blCn3Desc") or ""
    if not bl_cn1_desc.strip():
        try:
            if str(bl_cn3).strip().lstrip("0") == "4":
                bl_cn1_desc = bl_cn3_desc
        except Exception:
            pass

    # If this record is in GEOCODED_APPROX, use the cached town/county for display.
    # The ArcGIS schoolCode can be miscoded (e.g. Barnet property with schoolCode
    # pointing to Fair Haven), so the geocoded cache's propLocCty-derived values
    # are more accurate for the popup header.
    oid_str       = str(attr.get("OBJECTID") or "")
    geo_entry     = GEOCODED_APPROX.get(oid_str)
    if geo_entry and geo_entry.get("lat") and geo_entry.get("lon"):
        geo_city       = geo_entry.get("city") or loc_info["trustedTown"]
        geo_city_title = (geo_city or "").strip().title()
        geo_county     = TOWN_TO_COUNTY.get(geo_city_title)
        geo_county_code= str(geo_county).zfill(2) if geo_county else loc_info["trustedCountyCode"]
        geo_county_name= COUNTY_CODE_TO_NAME.get(geo_county_code) or loc_info["trustedCountyName"]
        display_town   = geo_city_title
        display_county_code = geo_county_code
        display_county_name = geo_county_name
    else:
        display_town        = loc_info["trustedTown"]
        display_county_code = loc_info["trustedCountyCode"]
        display_county_name = loc_info["trustedCountyName"]

    record = {
        # Location (trusted)
        "trustedTown":        display_town,
        "trustedCountyCode":  display_county_code,
        "trustedCountyName":  display_county_name,
        "correctedCounty":    loc_info["correctedCounty"],

        # Map display
        # resolve_coordinates validates the ArcGIS coords against the county
        # bounding box. If they're wrong it falls back to the town centroid
        # and sets approxLocation=True so the front-end can render those
        # markers differently (hollow circle with a note in the popup).
        "id":      attr.get("OBJECTID"),
        "address": attr.get("propLocStr"),
        "city":    loc_info["displayCity"],
        "price":   attr.get("ValPdOrTrn") or 0,
        "date":    attr.get("closeDate"),
        **dict(zip(
            ("lat", "lon", "approxLocation", "isCentroid"),
            resolve_coordinates(
                attr.get("Latitude"), attr.get("Longitude"),
                loc_info["trustedTown"],
                object_id=attr.get("OBJECTID"),
                match_method=attr.get("MatchMthod")
            )
        )),

        # Export columns — using actual abbreviated field names from ArcGIS service
        "sellerEntityName":            attr.get("sellEntNam"),
        "sellerLastName":              attr.get("sellLstNam"),
        "sellerFirstName":             attr.get("sellFstNam"),
        "buyerEntityName":             attr.get("buyEntNam"),
        "buyerLastName":               attr.get("buyLstNam"),
        "buyerFirstName":              attr.get("buyFstNam"),
        "propertyLocationStreet":      attr.get("propLocStr"),
        "propertyLocationCity":        attr.get("propLocCty"),
        "countyCode":                  loc_info["rawCountyCode"],
        "landSize":                    attr.get("landSize"),
        "span":                        loc_info["span"],
        "townCode":                    loc_info["townCode"],
        "schoolCode":                  loc_info["schoolCode"],
        "closingDate":                 attr.get("closeDate"),
        "dateSellerAcquired":          attr.get("SellerAcq"),
        "propertyTaxExemption":        attr.get("prTxEx"),
        "propertyTaxExemptionDesc":    attr.get("prTxExDesc"),
        "familyMemberDesc":            attr.get("famMemDesc"),
        "LGTExemption":                attr.get("LGTEx"),
        "LGTExemptionDesc":            attr.get("LGTExDesc"),
        "interestPropertyType":        attr.get("intPrpType"),
        "interestUndivPercentDesc":    attr.get("intUDPdesc"),
        "interestUndivPercent":        attr.get("intUDP"),
        "interestPropertyTypeOther":   attr.get("intPrTypOt"),
        "sellerAcquire":               attr.get("sellAq"),
        "sellerAcquireDesc":           attr.get("sellAqDesc"),
        "sellerAcquireOther":          attr.get("sellAqOthr"),
        "familyMember":                attr.get("famMem"),
        "buildingConstruction1":       attr.get("blCn1"),
        "buildingConstruction1Desc":   bl_cn1_desc,
        "buildingConstruction2":       attr.get("blCn2"),
        "buildingConstruction2Desc":   attr.get("blCn2Desc"),
        "buildingConstruction3":       attr.get("blCn3"),
        "buildingConstruction3Desc":   attr.get("blCn3Desc"),
        "buildingConstructionUnits05": attr.get("blCnUnts05"),
        "buildingConstruction20":      attr.get("blConstr20"),
        "buildingConstructionDwellingUnits06": attr.get("bCnDUs06"),
        "tenantPurchase":              attr.get("tenantPrch"),
        "financing":                   attr.get("financing"),
        "enrolledCurrentUse":          attr.get("enrCrntUse"),
        "currentUseEnrollmentContinue": attr.get("cUseEnCont"),
        "sellerUseOfProperty":         attr.get("sUsePr"),
        "sellerUseOfPropertyDesc":     attr.get("sUsePrDesc"),
        "sellerUseOfPropertyExplain":  attr.get("sUsePrExpl"),
        "buyerUseOfProperty":          attr.get("bUsePr"),
        "buyerUseOfPropertyDesc":      attr.get("bUsePrDesc"),
        "buyerUseOfPropertyExplain":   attr.get("bUsePrExpl"),
        "rentedBefore":                attr.get("rntdBefore"),
        "rentedAfter":                 attr.get("rntdAfter"),
        "developmentPrevConv":         attr.get("devPrevCnv"),
        "buyerAdjoiningProperty":      attr.get("buyrAdjPrp"),
        "ValuePaidOrTransferred":      attr.get("ValPdOrTrn"),
        "PersonalPropValuePaidOrTrans": attr.get("PrPrVlPdTr"),
        "RealPropValuePaidOrTrans":    attr.get("RlPrVlPdTr"),
        "TownBookNumber":              attr.get("TownBkNum"),
        "TownPageNumber":              attr.get("TownPgNum"),
        "TownCityorTown":              attr.get("TownCtyOrT"),
        "TownParcelIDNo":              attr.get("TownParcID"),
        "TownDateOfRecord":            attr.get("TownDteRec"),
        "TownGrandListCategory":       attr.get("TownGlCat"),
        "TownGrandListValue":          attr.get("TownGlValu"),
        "TownGrandListYear":           attr.get("TownGlYear"),
        "TownSpan":                    attr.get("TownSpan"),
        "TownSubdivision":             attr.get("TownSubdiv"),
        # Tax calculation fields
        "principalResidenceSRValue":   attr.get("prResSRVal"),
        "specialRateTaxDue":           attr.get("spRteTxDue"),
        "totalSpclRateDue":            attr.get("tlSpRteDue"),
        "ValueSubjecttoGenRate":       attr.get("VlSbjGnRte"),
        "GenRateTaxDue":               attr.get("GenRtTxDue"),
        "TotalTaxDue":                 attr.get("TotlTaxDue"),
        "exempt99Eligible":            attr.get("ex99Elig"),
        "exempt99TaxDue":              attr.get("ex99TxDue"),
        # Buyer/Seller addresses (used for PTT-172 only, not displayed in popup)
        "sellerStreet":                attr.get("sellerStrt"),
        "sellerCity":                  attr.get("sellerCity"),
        "sellerState":                 attr.get("sellerSt"),
        "sellerZip":                   attr.get("sellerZip"),
        "buyerStreet":                 attr.get("buyerStrt"),
        "buyerCity":                   attr.get("buyerCity"),
        "buyerState":                  attr.get("buyerState"),
        "buyerZip":                    attr.get("buyerZip"),
        # Export lat/lon: use resolved coords (centroid fallback if source coords are bad)
        "Latitude":                    resolve_coordinates(
                                           attr.get("Latitude"), attr.get("Longitude"),
                                           loc_info["rawCountyCode"],
                                           loc_info["trustedTown"])[0],
        "Longitude":                   resolve_coordinates(
                                           attr.get("Latitude"), attr.get("Longitude"),
                                           loc_info["rawCountyCode"],
                                           loc_info["trustedTown"])[1],
        "additionalSellerNames":       attr.get("addSellNam"),
        "additionalBuyerNames":        attr.get("addBuyrNam"),
    }

    if not apply_python_filters(record, filters):
        return None

    return record


def format_epoch_date(epoch_ms):
    """Convert epoch milliseconds to MM/DD/YYYY string."""
    if epoch_ms is None:
        return ""
    try:
        dt = datetime.fromtimestamp(int(epoch_ms) / 1000, tz=timezone.utc)
        return dt.strftime("%m/%d/%Y")
    except Exception:
        return str(epoch_ms)


def format_span(raw_span):
    """Format SPAN as XXX-XXX-XXXXX."""
    if raw_span is None:
        return ""
    s = str(int(raw_span)).zfill(11) if str(raw_span).isdigit() else str(raw_span)
    s = s.replace("-", "").zfill(11)
    if len(s) >= 11:
        return f"{s[0:3]}-{s[3:6]}-{s[6:11]}"
    return s


def format_boolean(val):
    """Convert ArcGIS string boolean ('TRUE'/'FALSE') to Yes / No."""
    if val is None or val == "":
        return "No"
    s = str(val).strip().upper()
    if s == "TRUE":
        return "Yes"
    return "No"  # 'FALSE', '0', blank, etc.


def format_interest_percent(val):
    """If zero, blank, or None → show 100%. Otherwise format as percentage."""
    if val is None or val == "" or val == 0:
        return "100%"
    try:
        f = float(val)
        if f == 0:
            return "100%"
        return f"{f:g}%"
    except Exception:
        return str(val)


def format_building_type(raw_desc):
    """Remove numeric code prefix from building type description.
    e.g. '02. Single Family Dwelling' → 'Single Family Dwelling'
    """
    if not raw_desc:
        return ""
    cleaned = re.sub(r"^\d+\.?\s*", "", str(raw_desc)).strip()
    return cleaned


def format_grand_list_category(code_val):
    """Look up Grand List Category description from numeric code."""
    if code_val is None or code_val == "":
        return ""
    try:
        code_str = str(int(float(str(code_val))))
        desc = GL_CATEGORY_LOOKUP.get(code_str) or GL_CATEGORY_LOOKUP.get(str(code_val))
        if desc:
            return desc
        return str(code_val)
    except Exception:
        return str(code_val)


def time_since_last_sale(epoch_ms):
    """Return 'X years Y months' between closeDate and today. No days."""
    if epoch_ms is None or epoch_ms == "":
        return ""
    try:
        close_dt = datetime.fromtimestamp(int(epoch_ms) / 1000, tz=timezone.utc)
        today    = datetime.now(tz=timezone.utc)

        years  = today.year  - close_dt.year
        months = today.month - close_dt.month

        if months < 0:
            years  -= 1
            months += 12

        parts = []
        if years > 0:
            parts.append(f"{years} year{'s' if years != 1 else ''}")
        if months > 0 or years == 0:
            parts.append(f"{months} month{'s' if months != 1 else ''}")
        return " ".join(parts)
    except Exception:
        return ""


# -----------------------------------------------------------------------
# Export column order and display names
# -----------------------------------------------------------------------
EXPORT_COLUMNS = [
    ("dateSellerAcquired",          "Date Seller Acquired"),
    ("timeSinceLastSale",           "Time Since Last Sale"),
    ("span",                        "SPAN"),
    ("TownParcelIDNo",              "Town Parcel ID"),
    ("closingDate",                 "Closing Date"),
    ("ValuePaidOrTransferred",      "Price Paid"),
    ("PersonalPropValuePaidOrTrans","Personal Prop Val"),
    ("RealPropValuePaidOrTrans",    "Real Property Value"),
    ("propertyLocationStreet",      "Property Street"),
    ("propertyLocationCity",        "City"),
    ("landSize",                    "Land Size (acres)"),
    ("buildingConstruction1Desc",   "Building Type"),
    ("buildingConstructionDwellingUnits06", "Dwelling Units"),
    ("buildingConstruction20",      "Other Uses"),
    ("sellerUseOfPropertyDesc",     "Seller Use of Property"),
    ("buyerUseOfPropertyDesc",      "Buyer Use of Property"),
    ("interestUndivPercentDesc",    "Interest Type"),
    ("buyerLastName",               "Buyer Last Name"),
    ("buyerFirstName",              "Buyer First Name"),
    ("buyerEntityName",             "Buyer Entity Name"),
    ("sellerLastName",              "Seller Last Name"),
    ("sellerFirstName",             "Seller First Name"),
    ("sellerEntityName",            "Seller Entity Name"),
    ("trustedCountyName",           "County"),
    ("TownCityorTown",              "Property City"),
    ("trustedTown",                 "Town"),
    ("interestUndivPercent",        "Interest Percent"),
    ("propertyTaxExemptionDesc",    "PTT Exemption"),
    ("LGTExemptionDesc",            "LGT Exemption"),
    ("familyMemberDesc",            "Family Member Relationship"),
    ("financing",                   "Financing"),
    ("enrolledCurrentUse",          "Enrolled Current Use"),
    ("currentUseEnrollmentContinue","Current Use Continues"),
    ("tenantPurchase",              "Tenant Purchase"),
    ("rentedBefore",                "Rented Before"),
    ("rentedAfter",                 "Rented After"),
    ("developmentPrevConv",         "Development Rights Previously Conveyed"),
    ("buyerAdjoiningProperty",      "Buyer Owns Adjoining Property"),
    ("TownBookNumber",              "Town Book Number"),
    ("TownPageNumber",              "Town Page Number"),
    ("TownDateOfRecord",            "Town Date of Record"),
    ("TownGrandListCategory",       "Grand List Category"),
    ("TownSubdivision",             "Subdivision"),
    ("Latitude",                    "Latitude"),
    ("Longitude",                   "Longitude"),
    ("schoolCode",                  "School Code"),
    ("townCode",                    "Town Code"),
    ("additionalSellerNames",       "Other Sellers"),
    ("additionalBuyerNames",        "Other Buyers"),
]

# Fields that store boolean values and should display as Yes/No
BOOLEAN_FIELDS = {
    "enrolledCurrentUse",
    "currentUseEnrollmentContinue",
    "tenantPurchase",
    "rentedBefore",
    "rentedAfter",
    "developmentPrevConv",
    "buyerAdjoiningProperty",
    "TownSubdivision",
}

# Date fields (epoch ms → MM/DD/YYYY)
DATE_FIELDS = {"closingDate", "dateSellerAcquired", "TownDateOfRecord"}


# Date column labels that should be written as real Excel dates
DATE_EXPORT_LABELS = {lbl for key, lbl in EXPORT_COLUMNS
                      if key in DATE_FIELDS}

def _parse_export_date(val):
    """Convert MM/DD/YYYY string to datetime.date for Excel. Returns None if unparseable."""
    if not val or not isinstance(val, str):
        return None
    try:
        from datetime import date as _date
        m, d, y = val.split("/")
        return _date(int(y), int(m), int(d))
    except Exception:
        return None


def format_record_for_export(rec):
    """Format a record dict for CSV/XLSX export."""
    out = {}

    # Pre-compute Time Since Last Sale from the raw closingDate epoch
    time_since = time_since_last_sale(rec.get("closingDate"))

    for key, label in EXPORT_COLUMNS:
        if key == "timeSinceLastSale":
            out[label] = time_since
            continue

        val = rec.get(key, "")

        if key in DATE_FIELDS:
            val = format_epoch_date(val)
        elif key == "span":
            val = format_span(val)
        elif key in BOOLEAN_FIELDS:
            val = format_boolean(val)
        elif key == "TownGrandListCategory":
            val = format_grand_list_category(val)
        elif key == "buildingConstruction1Desc":
            val = format_building_type(val)
        elif key == "interestUndivPercent":
            val = format_interest_percent(val)
        elif val is None:
            val = ""

        out[label] = val

    return out


# ==============================
# ROUTES
# ==============================

# alias so auth.py's url_for("index") resolves correctly
@app.route("/")
@login_required
def index():
    return home()

def home():
    # Admin bypass — only active when ADMIN_PASSWORD env var is set (local dev only)
    if ADMIN_PASSWORD_SET and session.get("user_email", "").lower() == ADMIN_EMAIL_LOCAL:
        return render_template("map.html",
                               user_email=session["user_email"],
                               subscription_status="active",
                               trial_days=0)
    user = db_fetchone(_q("SELECT * FROM users WHERE id = ?"), (session["user_id"],))
    trial_days = days_left_in_trial(user) if user["subscription_status"] in ("trial", "trialing") else 0
    return render_template("map.html",
                           user_email=user["email"],
                           subscription_status=user["subscription_status"],
                           trial_days=trial_days)


@app.route("/codes")
def codes():
    """Return all code lookup tables to the frontend."""
    return jsonify({
        "ptt_exemptions":          VT_CODES.get("ptt_exemptions", {}),
        "family_member_codes":      VT_CODES.get("family_member_codes", {}),
        "land_gains_exemptions":    VT_CODES.get("land_gains_exemptions", {}),
        "how_acquired_codes":       VT_CODES.get("how_acquired_codes", {}),
        "interest_types":           VT_CODES.get("interest_types", {}),
        "building_types":           VT_CODES.get("building_types", {}),
        "use_of_property":          VT_CODES.get("use_of_property", {}),
        "grand_list_categories":    VT_CODES.get("grand_list_categories", {}),
        "withholding_exemption_codes": VT_CODES.get("withholding_exemption_codes", {}),
        "town_centroids":           VT_CODES.get("town_centroids", {}),
    })


@app.route("/data")
def data():
    """Map data endpoint — viewport-bounded, returns up to 2000 records.

    When a name/SPAN/street filter is present the geometry bbox is dropped so
    ArcGIS searches the entire dataset by SQL.  Python then filters by name
    across all returned records, ensuring statewide name searches work.
    """
    xmin = request.args.get("xmin")
    ymin = request.args.get("ymin")
    xmax = request.args.get("xmax")
    ymax = request.args.get("ymax")

    filters = parse_filters(request.args)
    where   = build_where_clause(filters)

    # If any name / SPAN / street filter is active, skip the geometry bbox.
    # ArcGIS would apply the bbox first and return at most 2000 random records
    # from that area; Python name-matching then only sees that small pool.
    # Without the bbox ArcGIS returns up to 2000 records matching the SQL
    # filters (date, price, etc.) from the whole state, giving Python a much
    # larger and relevant pool to name-filter against.
    name_fields = [
        filters.get("seller_entity"), filters.get("seller_last"),
        filters.get("seller_first"),  filters.get("buyer_entity"),
        filters.get("buyer_last"),    filters.get("buyer_first"),
        filters.get("street"),        filters.get("span"),
    ]
    has_name_filter   = any(v for v in name_fields)
    has_town_filter   = bool(filters.get("towns",    "").strip())
    has_county_filter = bool(filters.get("counties", "").strip())

    if has_name_filter:
        # Name/SPAN/street search: no bbox — must search statewide to find
        # the record regardless of where it is on the map.
        features = fetch_all_features(where)

    elif has_county_filter:
        # County filter: use a generous county bbox instead of the viewport.
        # This returns all records within the county in one fast ArcGIS call
        # (no pagination) while still being far cheaper than a statewide fetch.
        # The county bbox is padded ~10 miles so border towns aren't clipped.
        # Misgeocoded records (coordinates outside bbox) are rare and are caught
        # by the post-fetch trustedCounty cross-check.
        county_codes = [c.strip().zfill(2) for c in filters["counties"].split(",") if c.strip()]
        # Union the bboxes of all selected counties
        lat_mins, lat_maxs, lon_mins, lon_maxs = [], [], [], []
        for code in county_codes:
            bounds = COUNTY_BOUNDS.get(code)
            if bounds:
                lat_min, lat_max, lon_min, lon_max = bounds
                lat_mins.append(lat_min); lat_maxs.append(lat_max)
                lon_mins.append(lon_min); lon_maxs.append(lon_max)
        if lat_mins:
            bbox_geo = {
                "geometry":     f"{min(lon_mins)},{min(lat_mins)},{max(lon_maxs)},{max(lat_maxs)}",
                "geometryType": "esriGeometryEnvelope",
                "inSR":         "4326",
                "spatialRel":   "esriSpatialRelIntersects",
            }
            features = fetch_features(where, bbox_geo, max_records=2000)
        else:
            # Unknown county code — fall back to statewide
            features = fetch_all_features(where)

    elif has_town_filter:
        # Town filter: build a bbox from the union of selected town centroids
        # with a generous ~0.15 degree (~10 mile) pad in each direction.
        # Single town fetches are fast — a town never exceeds ~500 records.
        town_names = [t.strip() for t in filters["towns"].split(",") if t.strip()]
        pad = 0.15
        t_lat_mins, t_lat_maxs, t_lon_mins, t_lon_maxs = [], [], [], []
        for town in town_names:
            centroid = TOWN_CENTROIDS.get(town.upper())
            if centroid:
                t_lat_mins.append(centroid["lat"] - pad)
                t_lat_maxs.append(centroid["lat"] + pad)
                t_lon_mins.append(centroid["lon"] - pad)
                t_lon_maxs.append(centroid["lon"] + pad)
        if t_lat_mins:
            town_bbox_geo = {
                "geometry":     f"{min(t_lon_mins)},{min(t_lat_mins)},{max(t_lon_maxs)},{max(t_lat_maxs)}",
                "geometryType": "esriGeometryEnvelope",
                "inSR":         "4326",
                "spatialRel":   "esriSpatialRelIntersects",
            }
            features = fetch_features(where, town_bbox_geo, max_records=2000)
        else:
            features = fetch_features(where, None, max_records=2000)

    else:
        geo_params = {
            "geometry":     f"{xmin},{ymin},{xmax},{ymax}",
            "geometryType": "esriGeometryEnvelope",
            "inSR":         "4326",
            "spatialRel":   "esriSpatialRelIntersects",
        }
        features = fetch_features(where, geo_params, max_records=2000)

    # Build a set of requested county codes for post-fetch validation.
    # Use trustedCountyCode (derived from schoolCode) — it is the most reliable
    # county indicator. countyCode on the form is self-reported and often wrong.
    # TOWNNAME is assigned by VCGI from GPS coordinates but is also unreliable
    # (e.g. 1543 Potter Hill Rd Readsboro has TOWNNAME=Rutland City).
    requested_counties = set(filters["counties"].split(",")) if filters["counties"] else set()

    results = []
    for f in features:
        rec = feature_to_record(f, filters)
        if rec is None:
            continue
        # County cross-check: use trustedCountyCode from schoolCode lookup.
        # If trustedCountyCode is null (junk schoolCode like 0 or 999),
        # fall back to propLocCty town name lookup before deciding.
        # If still null and a county filter is active, EXCLUDE the record —
        # don't let unverifiable records leak through county filters.
        if requested_counties:
            actual_county = rec.get("trustedCountyCode")
            if not actual_county:
                # Try propLocCty as a fallback county source
                prop_city = (rec.get("city") or "").strip().title()
                fallback = TOWN_TO_COUNTY.get(prop_city)
                actual_county = str(fallback).zfill(2) if fallback else None
            if not actual_county:
                continue  # Can't verify county — exclude from filtered results
            if str(actual_county).zfill(2) not in requested_counties:
                continue  # Wrong county — drop
        results.append({
                "id":               rec["id"],
                "address":          rec["address"],
                "city":             rec["city"],
                "price":            rec["ValuePaidOrTransferred"],
                "date":             rec["closingDate"],
                "lat":              rec["lat"],
                "lon":              rec["lon"],
                "trustedCountyCode": rec["trustedCountyCode"],
                "trustedCountyName": rec["trustedCountyName"],
                "trustedTown":       rec["trustedTown"],
                "schoolCode":        rec["schoolCode"],
                "span":              rec["span"],
                "correctedCounty":   rec["correctedCounty"],
                # Extra fields for popup display and tooltips
                "interestUndivPercentDesc": rec["interestUndivPercentDesc"],
                "buildingConstruction1Desc": rec["buildingConstruction1Desc"],
                "sellerUseOfPropertyDesc": rec["sellerUseOfPropertyDesc"],
                "buyerUseOfPropertyDesc":  rec["buyerUseOfPropertyDesc"],
                # Code fields for tooltip display
                "propertyTaxExemption":    rec["propertyTaxExemption"],
                "propertyTaxExemptionDesc": rec["propertyTaxExemptionDesc"],
                "familyMember":            rec["familyMember"],
                "familyMemberDesc":        rec["familyMemberDesc"],
                "LGTExemption":            rec["LGTExemption"],
                "LGTExemptionDesc":        rec["LGTExemptionDesc"],
                "sellerAcquire":           rec["sellerAcquire"],
                "sellerAcquireDesc":       rec["sellerAcquireDesc"],
                "interestPropertyType":    rec["interestPropertyType"],
                "buildingConstruction1":   rec["buildingConstruction1"],
                "buildingConstruction2":   rec["buildingConstruction2"],
                "buildingConstruction2Desc": rec["buildingConstruction2Desc"],
                "buildingConstruction3":   rec["buildingConstruction3"],
                "buildingConstruction3Desc": rec["buildingConstruction3Desc"],
                "sellerUseOfProperty":     rec["sellerUseOfProperty"],
                "buyerUseOfProperty":      rec["buyerUseOfProperty"],
                "approxLocation":          rec["approxLocation"],
                "isCentroid":              rec.get("isCentroid", rec["approxLocation"]),
                "sellerLastName":   rec["sellerLastName"],
                "sellerFirstName":  rec["sellerFirstName"],
                "sellerEntityName": rec["sellerEntityName"],
                "buyerLastName":    rec["buyerLastName"],
                "buyerFirstName":   rec["buyerFirstName"],
                "buyerEntityName":  rec["buyerEntityName"],
            })

    return jsonify({"data": results})


@app.route("/data/approx/all")
@login_required
def data_approx_all():
    """Serve pre-geocoded approx records from geocoded_approx.json.

    Built from geocode_approx.py, updated weekly after ArcGIS refreshes.
    Returns only records with valid coordinates (null-coord entries excluded).
    Frontend caches this once at page load and filters client-side on pan/zoom
    — no ArcGIS call needed, loads in under a second.
    """
    results = []
    for oid, entry in GEOCODED_APPROX.items():
        lat = entry.get("lat")
        lon = entry.get("lon")

        if lat is not None and lon is not None:
            # Successfully geocoded via SPAN/Nominatim.
            # Serve these here so records with null/bad ArcGIS geometry still appear.
            # If /data also returns the record, both land at identical coords and
            # merge into one popup via markerRefs grouping.
            is_centroid_flag = False
        else:
            # Null-coord record — couldn't be geocoded.
            # Fall back to town centroid so it shows as an orange circle.
            trusted_town = entry.get("trustedTown") or entry.get("city") or ""
            centroid = TOWN_CENTROIDS.get(trusted_town.strip().upper())
            if not centroid:
                continue  # no centroid available — skip
            lat = centroid["lat"]
            lon = centroid["lon"]
            is_centroid_flag = True
        # Use propLocCty (city field) for popup header — more accurate than
        # trustedTown which comes from schoolCode and can be miscoded.
        raw_city = (entry.get("city") or entry.get("trustedTown") or "").strip()
        # Strip trailing state abbreviations like "Barnet Vt" -> "Barnet"
        city = re.sub(r'\s+(VT|Vermont|Vt)\.?$', '', raw_city, flags=re.IGNORECASE).strip()

        # Derive county from propLocCty town name for accurate county display.
        city_upper   = city.strip().upper()
        city_county  = TOWN_TO_COUNTY.get(city.strip().title()) or TOWN_TO_COUNTY.get(city_upper.title())
        county_code2 = str(city_county).zfill(2) if city_county else entry.get("trustedCountyCode")
        county_name  = COUNTY_CODE_TO_NAME.get(county_code2) or entry.get("trustedCountyName") or ""

        results.append({
            "id":      int(oid),
            "address": entry.get("address", ""),
            "city":    city,
            "lat":     lat,
            "lon":     lon,
            "approxLocation": True,
            "isCentroid":     is_centroid_flag,
            # Minimal fields — popup will fetch full details on click via /ptt172
            "price":   None,
            "date":    None,
            "trustedCountyCode": county_code2 or entry.get("trustedCountyCode"),
            "trustedCountyName": county_name,
            "trustedTown":       city,  # use propLocCty for display
            "span":    entry.get("span"),
            "schoolCode": None,
            "correctedCounty": False,
            "interestUndivPercentDesc": None,
            "buildingConstruction1Desc": None,
            "sellerUseOfPropertyDesc": None,
            "buyerUseOfPropertyDesc": None,
            "propertyTaxExemption": None,
            "propertyTaxExemptionDesc": None,
            "familyMember": None,
            "familyMemberDesc": None,
            "LGTExemption": None,
            "LGTExemptionDesc": None,
            "sellerAcquire": None,
            "sellerAcquireDesc": None,
            "interestPropertyType": None,
            "buildingConstruction1": None,
            "buildingConstruction2": None,
            "buildingConstruction2Desc": None,
            "buildingConstruction3": None,
            "buildingConstruction3Desc": None,
            "sellerUseOfProperty": None,
            "buyerUseOfProperty": None,
            "sellerLastName": None,
            "sellerFirstName": None,
            "sellerEntityName": None,
            "buyerLastName": None,
            "buyerFirstName": None,
            "buyerEntityName": None,
        })
    return jsonify({"data": results})


@app.route("/data/approx/enrich")
@login_required
def data_approx_enrich():
    """Fetch full sale data for a single approx record by OBJECTID.
    Used to enrich merged approx records with price/date/building info.
    """
    obj_id = request.args.get("id", "")
    if not obj_id:
        return jsonify({"sale": None})
    try:
        obj_id_int = int(obj_id)
    except (TypeError, ValueError):
        return jsonify({"sale": None})

    features = fetch_features(f"OBJECTID={obj_id_int}", max_records=1)
    if not features:
        return jsonify({"sale": None})

    empty_filters = {
        "counties": "", "towns": "", "date_from": "", "date_to": "",
        "price_low": "", "price_high": "", "land_low": "", "land_high": "",
        "building_types": "", "interest_types": "", "seller_use": "",
        "buyer_use": "", "grand_list": "", "ptt_exemptions": "",
        "seller_entity": "", "seller_last": "", "seller_first": "",
        "buyer_entity": "", "buyer_last": "", "buyer_first": "",
        "street": "", "span": "",
        "enrolled_current_use": "", "dev_prev_conv": "",
        "buyer_adjoining": "", "foreclosed": "",
    }
    rec = feature_to_record(features[0], empty_filters)
    if not rec:
        return jsonify({"sale": None})

    return jsonify({"sale": {
        "price":                     rec["ValuePaidOrTransferred"],
        "date":                      rec["closingDate"],
        "buildingConstruction1Desc": rec["buildingConstruction1Desc"],
        "buildingConstruction2Desc": rec["buildingConstruction2Desc"],
        "buildingConstruction3Desc": rec["buildingConstruction3Desc"],
        "interestUndivPercentDesc":  rec["interestUndivPercentDesc"],
        "sellerLastName":            rec["sellerLastName"],
        "sellerFirstName":           rec["sellerFirstName"],
        "buyerLastName":             rec["buyerLastName"],
        "buyerFirstName":            rec["buyerFirstName"],
    }})


@app.route("/data/approx")
@login_required
def data_approx():
    """Return only approxLocation=True records matching the current filters.

    No viewport bbox — always searches the full dataset so centroid-placed
    markers are never missed due to pan/zoom position. The front-end keeps
    these in a separate layer that only refreshes on filter changes, not on
    every pan/zoom.
    """
    filters = parse_filters(request.args)
    where   = build_where_clause(filters)
    requested_counties = set(filters["counties"].split(",")) if filters["counties"] else set()

    features = fetch_all_features(where)

    results = []
    for f in features:
        rec = feature_to_record(f, filters)
        if rec is None:
            continue
        # Only include approx (centroid-placed) records
        if not rec.get("approxLocation"):
            continue
        # County cross-check: use trustedCountyCode from schoolCode lookup
        if requested_counties:
            actual_county = rec.get("trustedCountyCode")
            if actual_county and str(actual_county).zfill(2) not in requested_counties:
                continue
        results.append({
            "id":               rec["id"],
            "address":          rec["address"],
            "city":             rec["city"],
            "price":            rec["ValuePaidOrTransferred"],
            "date":             rec["closingDate"],
            "lat":              rec["lat"],
            "lon":              rec["lon"],
            "trustedCountyCode": rec["trustedCountyCode"],
            "trustedCountyName": rec["trustedCountyName"],
            "trustedTown":       rec["trustedTown"],
            "schoolCode":        rec["schoolCode"],
            "span":              rec["span"],
            "correctedCounty":   rec["correctedCounty"],
            "interestUndivPercentDesc": rec["interestUndivPercentDesc"],
            "buildingConstruction1Desc": rec["buildingConstruction1Desc"],
            "sellerUseOfPropertyDesc": rec["sellerUseOfPropertyDesc"],
            "buyerUseOfPropertyDesc":  rec["buyerUseOfPropertyDesc"],
            "propertyTaxExemption":    rec["propertyTaxExemption"],
            "propertyTaxExemptionDesc": rec["propertyTaxExemptionDesc"],
            "familyMember":            rec["familyMember"],
            "familyMemberDesc":        rec["familyMemberDesc"],
            "LGTExemption":            rec["LGTExemption"],
            "LGTExemptionDesc":        rec["LGTExemptionDesc"],
            "sellerAcquire":           rec["sellerAcquire"],
            "sellerAcquireDesc":       rec["sellerAcquireDesc"],
            "interestPropertyType":    rec["interestPropertyType"],
            "buildingConstruction1":   rec["buildingConstruction1"],
            "buildingConstruction2":   rec["buildingConstruction2"],
            "buildingConstruction2Desc": rec["buildingConstruction2Desc"],
            "buildingConstruction3":   rec["buildingConstruction3"],
            "buildingConstruction3Desc": rec["buildingConstruction3Desc"],
            "sellerUseOfProperty":     rec["sellerUseOfProperty"],
            "buyerUseOfProperty":      rec["buyerUseOfProperty"],
            "approxLocation":          True,
            "sellerLastName":   rec["sellerLastName"],
            "sellerFirstName":  rec["sellerFirstName"],
            "sellerEntityName": rec["sellerEntityName"],
            "buyerLastName":    rec["buyerLastName"],
            "buyerFirstName":   rec["buyerFirstName"],
            "buyerEntityName":  rec["buyerEntityName"],
        })

    return jsonify({"data": results})


@app.route("/export")
def export():
    """
    Export endpoint — fetches ALL matching records statewide (no viewport).
    Returns XLSX by default, CSV if ?format=csv.
    Accepts same filter params as /data plus optional ?ids= for selected-only export.
    """
    fmt     = request.args.get("format", "xlsx").lower()
    ids_raw = request.args.get("ids", "")

    filters = parse_filters(request.args)
    where   = build_where_clause(filters)

    selected_ids = [i.strip() for i in ids_raw.split("|") if i.strip()] if ids_raw else []
    if selected_ids:
        ids_sql = ",".join(selected_ids)
        where = f"({where}) AND OBJECTID IN ({ids_sql})"
        features = fetch_features(where, max_records=len(selected_ids) + 10)
    else:
        features = fetch_all_features(where)

    records = []
    for f in features:
        rec = feature_to_record(f, filters)
        if rec is not None:
            records.append(format_record_for_export(rec))

    if not records:
        return jsonify({"error": "No records matched the current filters."}), 404

    # Sort by Closing Date ascending (convert MM/DD/YYYY -> sortable YYYY-MM-DD)
    closing_label = next((lbl for _, lbl in EXPORT_COLUMNS if lbl == "Closing Date"), "Closing Date")
    def _date_sort_key(r):
        v = r.get(closing_label) or ""
        try: parts = v.split("/"); return f"{parts[2]}-{parts[0]}-{parts[1]}"
        except: return v
    records.sort(key=_date_sort_key)

    headers = [label for _, label in EXPORT_COLUMNS]

    if fmt == "csv":
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(records)
        csv_data = output.getvalue()
        return Response(
            csv_data,
            mimetype="text/csv",
            headers={"Content-Disposition": "attachment; filename=vt_property_transfers.csv"}
        )
    else:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return jsonify({"error": "openpyxl not installed. Run: pip install openpyxl"}), 500

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "VT Property Transfers"

        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF")

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx, rec in enumerate(records, 2):
            for col_idx, header in enumerate(headers, 1):
                raw = rec.get(header, "")
                cell = ws.cell(row=row_idx, column=col_idx)
                if header in DATE_EXPORT_LABELS:
                    dt = _parse_export_date(raw)
                    if dt:
                        cell.value = dt
                        cell.number_format = "MM/DD/YYYY"
                    else:
                        cell.value = raw
                else:
                    cell.value = raw

        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

        ws.freeze_panes = "A2"

        xlsx_buffer = io.BytesIO()
        wb.save(xlsx_buffer)
        xlsx_buffer.seek(0)

        return Response(
            xlsx_buffer.read(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=vt_property_transfers.xlsx"}
        )


# ---------------------------------------------------------------------------
# Server-side map image generator
# Fetches OSM tiles for the given bounds/zoom, draws polygons + stats boxes.
# ---------------------------------------------------------------------------
def generate_map_image(map_meta, group_records):
    """
    map_meta: {
        bounds: {north, south, east, west},
        zoom: int,
        shapes: [{label, color, latlngs: [[lat,lng],...], stats: {...}}]
                 (circles sent as polygon approximation from frontend)
    }
    Returns PNG bytes or None.
    """
    import math, struct, zlib
    import urllib.request
    try:
        from PIL import Image, ImageDraw, ImageFont, ImageColor
    except ImportError:
        return None

    meta           = map_meta or {}
    bounds         = meta.get("bounds", {})
    zoom           = int(meta.get("zoom", 14))
    shapes         = meta.get("shapes", [])
    selected_dots  = meta.get("selected_dots",   [])  # [[lat,lng],...]
    unselected_dots= meta.get("unselected_dots",  [])  # [[lat,lng],...]

    north = float(bounds.get("north",  44.20))
    south = float(bounds.get("south",  44.08))
    east  = float(bounds.get("east",  -72.55))
    west  = float(bounds.get("west",  -72.75))

    # ---- Tile math ----
    def latlon_to_tile(lat, lon, z):
        n = 2 ** z
        x = int((lon + 180) / 360 * n)
        lat_r = math.radians(lat)
        y = int((1 - math.log(math.tan(lat_r) + 1/math.cos(lat_r)) / math.pi) / 2 * n)
        return x, y

    def tile_to_latlon(x, y, z):
        n = 2 ** z
        lon = x / n * 360 - 180
        lat_r = math.atan(math.sinh(math.pi * (1 - 2 * y / n)))
        return math.degrees(lat_r), lon

    tx_min, ty_min = latlon_to_tile(north, west, zoom)
    tx_max, ty_max = latlon_to_tile(south, east, zoom)
    tx_min, tx_max = min(tx_min, tx_max), max(tx_min, tx_max)
    ty_min, ty_max = min(ty_min, ty_max), max(ty_min, ty_max)

    TILE_SIZE = 256
    cols = tx_max - tx_min + 1
    rows = ty_max - ty_min + 1
    # Cap at reasonable size
    if cols > 8: tx_max = tx_min + 7; cols = 8
    if rows > 8: ty_max = ty_min + 7; rows = 8

    img_w = cols * TILE_SIZE
    img_h = rows * TILE_SIZE
    canvas = Image.new("RGB", (img_w, img_h), (240, 240, 240))

    headers = {"User-Agent": "VTPropertyTool/1.0 (appraiservt@gmail.com)"}
    for tx in range(tx_min, tx_max + 1):
        for ty in range(ty_min, ty_max + 1):
            url = f"https://tile.openstreetmap.org/{zoom}/{tx}/{ty}.png"
            try:
                req  = urllib.request.Request(url, headers=headers)
                with urllib.request.urlopen(req, timeout=5) as resp:
                    tile_data = resp.read()
                tile_img = Image.open(io.BytesIO(tile_data)).convert("RGB")
                px = (tx - tx_min) * TILE_SIZE
                py = (ty - ty_min) * TILE_SIZE
                canvas.paste(tile_img, (px, py))
            except Exception:
                pass  # leave grey if tile fails

    # ---- Coordinate conversion ----
    # Top-left of canvas corresponds to tile (tx_min, ty_min)
    origin_lat, origin_lon = tile_to_latlon(tx_min, ty_min, zoom)
    n_tiles_lat, n_tiles_lon = tile_to_latlon(tx_min + cols, ty_min + rows, zoom)

    def latlon_to_px(lat, lon):
        # Mercator pixel position
        n = 2 ** zoom
        px_x = (lon + 180) / 360 * n * TILE_SIZE - tx_min * TILE_SIZE
        lat_r = math.radians(lat)
        py_y  = (1 - math.log(math.tan(lat_r) + 1/math.cos(lat_r)) / math.pi) / 2 * n * TILE_SIZE - ty_min * TILE_SIZE
        return int(px_x), int(py_y)

    draw = ImageDraw.Draw(canvas, "RGBA")

    # Try to load a font, fall back to default
    try:
        fnt_bold  = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 14)
        fnt_small = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",     12)
        fnt_lbl   = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 13)
    except Exception:
        fnt_bold  = ImageFont.load_default()
        fnt_small = fnt_bold
        fnt_lbl   = fnt_bold

    PURPLE      = (123, 0, 212)
    PURPLE_FILL = (123, 0, 212, 40)   # semi-transparent
    WHITE       = (255, 255, 255)
    DARK        = (40,  40,  40)
    BORDER      = (123, 0, 212, 200)

    # ---- Draw each shape ----
    box_positions = {}  # shape label -> (box_x, box_y, box_w, box_h)
    anchor_positions = {}

    for i, shape in enumerate(shapes):
        latlngs = shape.get("latlngs", [])
        if not latlngs:
            continue
        label = shape.get("label", f"Area {i+1}")
        stats = shape.get("stats", {})

        pts = [latlon_to_px(ll[0], ll[1]) for ll in latlngs]
        if not pts:
            continue

        # Polygon fill + outline
        draw.polygon(pts, fill=(123, 0, 212, 35), outline=None)
        for j in range(len(pts)):
            draw.line([pts[j], pts[(j+1) % len(pts)]], fill=(60, 80, 120), width=2)

        # Centroid for anchor dot
        cx = int(sum(p[0] for p in pts) / len(pts))
        cy = int(sum(p[1] for p in pts) / len(pts))
        anchor_positions[label] = (cx, cy)
        draw.ellipse([cx-5, cy-5, cx+5, cy+5], fill=PURPLE, outline=WHITE)

        # Stats box position: 220px right, offset by row
        bx = min(cx + 220, img_w - 200)
        by = max(cy - 80 - i * 22, 10)
        bx = max(bx, 5)

        # Measure stats text
        fmt_m = lambda v: f"${int(v):,}" if v is not None else "N/A"
        rows_data = [
            ("Sales (>$0):", str(stats.get("count", 0))),
            ("High:",        fmt_m(stats.get("high"))),
            ("Low:",         fmt_m(stats.get("low"))),
            ("Average:",     fmt_m(stats.get("avg"))),
            ("Median:",      fmt_m(stats.get("median"))),
        ]
        box_w  = 210
        hdr_h  = 28
        row_h  = 20
        pad    = 8
        box_h  = hdr_h + pad + len(rows_data) * row_h + pad

        # Shadow
        shadow_off = 3
        draw.rounded_rectangle(
            [bx+shadow_off, by+shadow_off, bx+box_w+shadow_off, by+box_h+shadow_off],
            radius=6, fill=(0, 0, 0, 60)
        )

        # White body
        draw.rounded_rectangle([bx, by, bx+box_w, by+box_h], radius=6,
                                fill=WHITE, outline=PURPLE, width=2)
        # Purple header
        draw.rounded_rectangle([bx, by, bx+box_w, by+hdr_h], radius=6,
                                fill=PURPLE)
        # Cover bottom corners of header (make it flat on bottom)
        draw.rectangle([bx, by+hdr_h-6, bx+box_w, by+hdr_h], fill=PURPLE)

        # Header text
        draw.text((bx+10, by+7), label, font=fnt_lbl, fill=WHITE)

        # Stats rows
        for ri, (k, v) in enumerate(rows_data):
            ry = by + hdr_h + pad + ri * row_h
            draw.text((bx+8,       ry), k, font=fnt_small, fill=DARK)
            # Right-align value
            try:
                vw = fnt_small.getlength(v)
            except Exception:
                vw = len(v) * 7
            draw.text((bx+box_w-8-vw, ry), v,
                      font=fnt_bold if ri == 0 else fnt_small, fill=DARK)

        box_positions[label] = (bx, by, box_w, box_h)

    # ---- Draw tie lines (anchor dot -> box edge) ----
    for label, (bx, by, bw, bh) in box_positions.items():
        if label not in anchor_positions:
            continue
        ax, ay = anchor_positions[label]
        # Connect to left-center of box
        ex, ey = bx, by + bh // 2
        # Dashed line
        dx, dy = ex - ax, ey - ay
        dist   = max(1, math.hypot(dx, dy))
        dash, gap = 8, 5
        step = dash + gap
        d = 0
        while d < dist:
            t0 = d / dist
            t1 = min((d + dash) / dist, 1.0)
            x0, y0 = int(ax + dx*t0), int(ay + dy*t0)
            x1, y1 = int(ax + dx*t1), int(ay + dy*t1)
            draw.line([(x0,y0),(x1,y1)], fill=PURPLE, width=2)
            d += step

    # ---- Draw Area label pills ----
    for label, (ax, ay) in anchor_positions.items():
        pill_pad_x, pill_pad_y = 10, 5
        try:
            tw = fnt_lbl.getlength(label)
        except Exception:
            tw = len(label) * 8
        pw = int(tw) + pill_pad_x * 2
        ph = 26
        px0 = ax - pw // 2
        py0 = ay - 40  # above anchor
        draw.rounded_rectangle([px0, py0, px0+pw, py0+ph],
                                radius=ph//2, fill=PURPLE)
        draw.text((px0+pill_pad_x, py0+5), label, font=fnt_lbl, fill=WHITE)

    # ---- Draw property dots (on top of everything) ----
    DOT_R = 5
    BLUE  = (0, 102, 255)
    DOT_DARK = (26, 26, 26)

    for ll in unselected_dots:
        px, py = latlon_to_px(ll[0], ll[1])
        draw.ellipse([px-DOT_R, py-DOT_R, px+DOT_R, py+DOT_R],
                     fill=(255, 255, 255), outline=DOT_DARK, width=2)

    for ll in selected_dots:
        px, py = latlon_to_px(ll[0], ll[1])
        draw.ellipse([px-DOT_R, py-DOT_R, px+DOT_R, py+DOT_R],
                     fill=BLUE, outline=BLUE)

    # ---- Export to PNG bytes ----
    out = io.BytesIO()
    canvas.save(out, format="PNG", optimize=True)
    out.seek(0)
    return out.read()


@app.route("/export_grouped", methods=["GET", "POST"])
def export_grouped():
    """
    Grouped export.
    Accepts POST with JSON body: {format, filters, groups}
    or GET with ?groups=JSON&format=...
    """
    import json as _json

    if request.method == "POST":
        data = request.get_json(force=True, silent=True) or {}
        fmt        = data.get("format", "xlsx").lower()
        groups     = data.get("groups", [])
        map_meta   = data.get("map_meta",   None)  # {bounds, zoom, shapes}
        screenshot = data.get("screenshot", None)  # base64 PNG from getDisplayMedia
        # Build a fake args-like object from the filters dict
        filter_dict = data.get("filters", {})
        class _FakeArgs(dict):
            def get(self, k, default=""):
                return super().get(k, default)
        filters = parse_filters(_FakeArgs(filter_dict))
    else:
        fmt        = request.args.get("format", "xlsx").lower()
        groups_raw = request.args.get("groups", "[]")
        try:
            groups = _json.loads(groups_raw)
        except Exception:
            return jsonify({"error": "Invalid groups parameter"}), 400
        filters    = parse_filters(request.args)
        map_meta   = None
        screenshot = None

    if not groups:
        return jsonify({"error": "No groups provided"}), 400

    # Fetch records for each group
    group_records = []
    for g in groups:
        ids = [str(i).strip() for i in g.get("ids", []) if str(i).strip()]
        stats = g.get("stats", {})
        label = g.get("label", "Area")
        if not ids:
            group_records.append({"label": label, "records": [], "stats": stats})
            continue
        ids_sql  = ",".join(ids)
        where    = f"OBJECTID IN ({ids_sql})"
        features = fetch_features(where, max_records=len(ids) + 10)
        recs = []
        for f in features:
            rec = feature_to_record(f, filters)
            if rec is not None:
                recs.append(format_record_for_export(rec))
        # Sort each group by Closing Date ascending
        closing_label = next((lbl for _, lbl in EXPORT_COLUMNS if lbl == "Closing Date"), "Closing Date")
        def _grp_date_key(r):
            v = r.get(closing_label) or ""
            try: parts = v.split("/"); return f"{parts[2]}-{parts[0]}-{parts[1]}"
            except: return v
        recs.sort(key=_grp_date_key)
        group_records.append({"label": label, "records": recs, "stats": stats})

    headers = [label for _, label in EXPORT_COLUMNS]

    def fmt_money(val):
        if val is None:
            return "N/A"
        try:
            return f"${int(val):,}"
        except Exception:
            return str(val)

    if fmt == "csv":
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=headers, extrasaction="ignore")
        first = True
        for g in group_records:
            if not first:
                writer.writerow({h: "" for h in headers})
                writer.writerow({h: "" for h in headers})
                writer.writerow({h: "" for h in headers})
                writer.writerow({h: "" for h in headers})
            first = False
            # Group header
            writer.writerow({headers[0]: f"=== {g['label']} ===", **{h: "" for h in headers[1:]}})
            writer.writeheader()
            writer.writerows(g["records"])
            # Blank row before stats
            writer.writerow({h: "" for h in headers})
            # Stats row
            st = g["stats"]
            writer.writerow({h: "" for h in headers})
            writer.writerow({
                headers[0]: f"{g['label']} Summary",
                headers[1]: f"Sales (>$0): {st.get('count','N/A')}",
                headers[2]: f"High: {fmt_money(st.get('high'))}",
                headers[3]: f"Low: {fmt_money(st.get('low'))}",
                headers[4]: f"Average: {fmt_money(st.get('avg'))}",
                headers[5]: f"Median: {fmt_money(st.get('median'))}",
                **{h: "" for h in headers[6:]}
            })
        return Response(
            output.getvalue(),
            mimetype="text/csv",
            headers={"Content-Disposition": "attachment; filename=vt_property_transfers_grouped.csv"}
        )
    else:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            return jsonify({"error": "openpyxl not installed. Run: pip install openpyxl"}), 500

        wb  = openpyxl.Workbook()
        ws  = wb.active
        ws.title = "VT Property Transfers"

        # Styles
        hdr_fill  = PatternFill("solid", fgColor="1F4E79")
        hdr_font  = Font(bold=True, color="FFFFFF")
        grp_fill  = PatternFill("solid", fgColor="2E75B6")
        grp_font  = Font(bold=True, color="FFFFFF", size=12)
        stat_fill = PatternFill("solid", fgColor="D9E1F2")
        stat_font = Font(bold=True, color="1F4E79")
        center    = Alignment(horizontal="center")

        cur_row = 1

        for g_idx, g in enumerate(group_records):
            # ---- Group label row ----
            # Fill every cell in the row with the blue style so the bar
            # spans all columns. Put the label text in column 1 only.
            # (Avoiding merge_cells entirely — it reliably drops the value
            # in some openpyxl versions on Python 3.14.)
            for col_idx in range(1, len(headers) + 1):
                c = ws.cell(row=cur_row, column=col_idx,
                            value=g["label"] if col_idx == 1 else "")
                c.fill = grp_fill
                c.font = grp_font
                if col_idx == 1:
                    c.alignment = Alignment(horizontal="left", vertical="center")
            cur_row += 1

            # ---- Column header row ----
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=cur_row, column=col_idx, value=h)
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = center
            cur_row += 1

            # ---- Data rows ----
            for rec in g["records"]:
                for col_idx, h in enumerate(headers, 1):
                    raw  = rec.get(h, "")
                    cell = ws.cell(row=cur_row, column=col_idx)
                    if h in DATE_EXPORT_LABELS:
                        dt = _parse_export_date(raw)
                        if dt:
                            cell.value = dt
                            cell.number_format = "MM/DD/YYYY"
                        else:
                            cell.value = raw
                    else:
                        cell.value = raw
                cur_row += 1

            # ---- Blank row before stats (keeps stats out of sort range) ----
            cur_row += 1

            # ---- Stats row ----
            st = g["stats"]
            stats_labels = [
                f"Sales (>$0): {st.get('count', 'N/A')}",
                f"High: {fmt_money(st.get('high'))}",
                f"Low: {fmt_money(st.get('low'))}",
                f"Average: {fmt_money(st.get('avg'))}",
                f"Median: {fmt_money(st.get('median'))}",
            ]
            for col_idx, val in enumerate(stats_labels, 1):
                cell = ws.cell(row=cur_row, column=col_idx, value=val)
                cell.fill = stat_fill
                cell.font = stat_font
            cur_row += 1

            # ---- 4 blank spacer rows between groups ----
            if g_idx < len(group_records) - 1:
                cur_row += 4

        # Column widths (skip MergedCell objects which have no column_letter)
        for col in ws.columns:
            first = next((c for c in col if hasattr(c, 'column_letter')), None)
            if not first:
                continue
            max_len = max((len(str(cell.value or "")) for cell in col if hasattr(cell, 'value')), default=10)
            ws.column_dimensions[first.column_letter].width = min(max_len + 2, 40)

        ws.freeze_panes = "A3"  # freeze past group label + header

        # ---- Map tab: server-side rendered polygon map ----
        if map_meta:
            from openpyxl.drawing.image import Image as XLImage
            try:
                img_bytes = generate_map_image(map_meta, group_records)
                if img_bytes:
                    img_stream = io.BytesIO(img_bytes)
                    ws_map = wb.create_sheet(title="Map")
                    ws_map.sheet_view.showGridLines = False
                    xl_img = XLImage(img_stream)
                    xl_img.width  = 1200
                    xl_img.height = 800
                    ws_map.add_image(xl_img, "A1")
                    ws_map.column_dimensions["A"].width = 160
                    ws_map.row_dimensions[1].height     = 600
            except Exception:
                pass  # non-fatal

        # ---- Screenshot tab: exact screen capture from browser ----
        if screenshot:
            import base64 as _b64
            from openpyxl.drawing.image import Image as XLImage
            try:
                raw = _b64.b64decode(screenshot)
                ws_sc = wb.create_sheet(title="Screen Capture")
                ws_sc.sheet_view.showGridLines = False
                xl_sc = XLImage(io.BytesIO(raw))
                # Fit to ~1200 wide preserving aspect ratio
                from PIL import Image as _PILImage
                pil = _PILImage.open(io.BytesIO(raw))
                orig_w, orig_h = pil.size
                scale = 1200 / orig_w if orig_w > 1200 else 1.0
                xl_sc.width  = int(orig_w * scale)
                xl_sc.height = int(orig_h * scale)
                ws_sc.add_image(xl_sc, "A1")
                ws_sc.column_dimensions["A"].width = xl_sc.width / 7.5
                ws_sc.row_dimensions[1].height     = xl_sc.height * 0.75
            except Exception:
                pass  # non-fatal

        xlsx_buf = io.BytesIO()
        wb.save(xlsx_buf)
        xlsx_buf.seek(0)
        return Response(
            xlsx_buf.read(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=vt_property_transfers_grouped.xlsx"}
        )


@app.route("/history")
def history():
    """Return ALL sales for a given SPAN (no date filter) so the popup can
    show older sales above the dotted-line divider even when a date range
    filter is active on the main map."""
    span_raw = request.args.get("span", "").strip().replace("-", "")
    if not span_raw:
        return jsonify({"data": []})

    where = f"CAST(span AS VARCHAR(20)) LIKE '%{span_raw}%'"
    features = fetch_all_features(where)

    # Parse filters without date range so all sales are returned
    empty_filters = {k: "" for k in [
        "counties", "towns", "date_from", "date_to", "price_low", "price_high",
        "land_low", "land_high", "street", "interest", "building", "seller_use",
        "buyer_use", "ptt_exemption", "grand_list", "seller_entity", "seller_last",
        "seller_first", "buyer_entity", "buyer_last", "buyer_first", "span",
        "dev_prev_conv", "buyer_adjoining", "enrolled_current_use", "foreclosed",
    ]}

    results = []
    for f in features:
        rec = feature_to_record(f, empty_filters)
        if rec is not None:
            results.append({
                "id":              rec["id"],
                "address":         rec["address"],
                "price":           rec["ValuePaidOrTransferred"],
                "date":            rec["closingDate"],
                "lat":             rec["lat"],
                "lon":             rec["lon"],
                "sellerLastName":  rec["sellerLastName"],
                "sellerFirstName": rec["sellerFirstName"],
                "sellerEntityName":rec["sellerEntityName"],
                "buyerLastName":   rec["buyerLastName"],
                "buyerFirstName":  rec["buyerFirstName"],
                "buyerEntityName": rec["buyerEntityName"],
                "buildingConstruction1Desc": rec["buildingConstruction1Desc"],
                "interestUndivPercentDesc":  rec["interestUndivPercentDesc"],
                "approxLocation":  rec["approxLocation"],
                "isCentroid":      rec.get("isCentroid", rec["approxLocation"]),
            })

    return jsonify({"data": results})


@app.route("/ptt172/preview")
def ptt172_preview():
    """HTML preview of PTT-172 data with hover tooltips on code fields."""
    from datetime import datetime as _dt, timezone as _tz

    rec_id = request.args.get("id", "")
    if not rec_id:
        return jsonify({"error": "Missing id"}), 400

    features = fetch_features(f"OBJECTID = {rec_id}", max_records=1)
    if not features:
        return jsonify({"error": "Record not found"}), 404

    empty_f = {k: "" for k in [
        "counties","towns","date_from","date_to","price_low","price_high",
        "land_low","land_high","street","interest","building","seller_use",
        "buyer_use","span","buyer_entity","buyer_last","buyer_first",
        "seller_entity","seller_last","seller_first",
    ]}
    rec = feature_to_record(features[0], empty_f)
    if not rec:
        return jsonify({"error": "Could not parse record"}), 500

    # Date formatting helpers
    def fmtd(epoch_ms):
        if not epoch_ms: return ""
        try: return _dt.fromtimestamp(int(epoch_ms)/1000, tz=_tz.utc).strftime("%m/%d/%Y")
        except: return ""

    def fmt_span(s):
        if not s: return ""
        s = str(s).replace("-","").zfill(11)
        return f"{s[:3]}-{s[3:6]}-{s[6:]}"

    # Time held
    def time_held(acq_ms, close_ms):
        if not acq_ms or not close_ms: return 0, 0
        try:
            d1 = _dt.fromtimestamp(int(acq_ms)/1000, tz=_tz.utc)
            d2 = _dt.fromtimestamp(int(close_ms)/1000, tz=_tz.utc)
            mt = (d2.year-d1.year)*12 + (d2.month-d1.month)
            if d2.day < d1.day: mt -= 1
            return mt//12, mt%12
        except: return 0, 0

    held_y, held_m = time_held(rec.get("dateSellerAcquired"), rec.get("closingDate"))

    # Building rows
    building_rows = []
    for i, (code_key, desc_key) in enumerate([
        ("buildingConstruction1", "buildingConstruction1Desc"),
        ("buildingConstruction2", "buildingConstruction2Desc"),
        ("buildingConstruction3", "buildingConstruction3Desc"),
    ], 1):
        c = rec.get(code_key)
        d = rec.get(desc_key) or BUILDING_TYPE_LOOKUP.get(str(c), "")
        if c:
            building_rows.append((c, d, i))

    from markupsafe import Markup

    # Pre-compute all code chip HTML so template needs no Python logic
    def mk_chip(val, table, label):
        if val is None or str(val).strip() == "": return Markup("")
        try: code_str = str(int(float(str(val)))).zfill(2)
        except: code_str = str(val)
        desc = VT_CODES.get(table, {}).get(str(int(float(str(val)))), "") if val else ""
        tip  = f"{label}: {code_str} \u2014 {desc}" if desc else f"{label}: {code_str}"
        tip  = tip.replace('"', '&quot;').replace("'", "&#39;")
        return Markup(f'<span class="code-chip" data-tip="{tip}">{code_str}</span>')

    def code_desc(val, table):
        if val is None or str(val).strip() == "": return ""
        try: key = str(int(float(str(val))))
        except: key = str(val)
        return VT_CODES.get(table, {}).get(key, "")

    chips = {
        "e1": mk_chip(rec.get("propertyTaxExemption"),  "ptt_exemptions",       "E1 PTT Exemption"),
        "e2": mk_chip(rec.get("familyMember"),           "family_member_codes",  "E2 Family Member"),
        "e3": mk_chip(rec.get("LGTExemption"),           "land_gains_exemptions","E3 Land Gains"),
        "f1": mk_chip(rec.get("sellerAcquire"),          "how_acquired_codes",   "F1 How Acquired"),
        "f2": mk_chip(rec.get("interestPropertyType"),   "interest_types",       "F2 Interest Type"),
        "h1": mk_chip(rec.get("sellerUseOfProperty"),    "use_of_property",      "H1 Seller Use"),
        "h2": mk_chip(rec.get("buyerUseOfProperty"),     "use_of_property",      "H2 Buyer Use"),
    }
    descs = {
        "e1": code_desc(rec.get("propertyTaxExemption"),  "ptt_exemptions")       or rec.get("propertyTaxExemptionDesc",""),
        "e2": code_desc(rec.get("familyMember"),           "family_member_codes")  or rec.get("familyMemberDesc",""),
        "e3": code_desc(rec.get("LGTExemption"),           "land_gains_exemptions") or rec.get("LGTExemptionDesc",""),
        "f1": code_desc(rec.get("sellerAcquire"),          "how_acquired_codes")   or rec.get("sellerAcquireDesc",""),
        "f2": code_desc(rec.get("interestPropertyType"),   "interest_types")       or rec.get("interestUndivPercentDesc",""),
        "h1": code_desc(rec.get("sellerUseOfProperty"),    "use_of_property")      or rec.get("sellerUseOfPropertyDesc",""),
        "h2": code_desc(rec.get("buyerUseOfProperty"),     "use_of_property")      or rec.get("buyerUseOfPropertyDesc",""),
    }
    # Building chips
    bldg_chips = []
    for i, (ck, dk) in enumerate([
        ("buildingConstruction1","buildingConstruction1Desc"),
        ("buildingConstruction2","buildingConstruction2Desc"),
        ("buildingConstruction3","buildingConstruction3Desc"),
    ], 1):
        c = rec.get(ck)
        if c:
            d = code_desc(c, "building_types") or rec.get(dk, "")
            bldg_chips.append((mk_chip(c, "building_types", f"F3 Building Type"), d, i))

    return render_template(
        "ptt172_preview.html",
        rec=rec,
        rec_id=rec_id,
        chips=chips,
        descs=descs,
        bldg_chips=bldg_chips,
        closing_date=fmtd(rec.get("closingDate")),
        acquired_date=fmtd(rec.get("dateSellerAcquired")),
        held_years=held_y,
        held_months=held_m,
        span_fmt=fmt_span(rec.get("span")),
    )


@app.route("/ptt172")
@app.route("/ptt172/<filename>")
def ptt172(filename=None):  # noqa: C901
    """Pre-filled PTT-172 PDF. Opens inline in browser by default.
    ?id=OBJECTID  &download=1 to force file download.
    The optional <filename> path segment (e.g. PTT-172_xxx.pdf) makes
    browsers use the correct .pdf extension when saving.
    """
    import io as _io
    from datetime import datetime as _dt, timezone as _tz
    try:
        from pypdf import PdfReader, PdfWriter
        from pypdf.generic import NameObject
        from reportlab.pdfgen import canvas as rl_canvas
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.units import inch
        from pypdf import PdfReader as _PR2
    except ImportError as _ie:
        return jsonify({"error": f"Missing library: {_ie}"}), 500

    rec_id   = request.args.get("id", "")
    download = request.args.get("download", "0") == "1"
    if not rec_id:
        return jsonify({"error": "Missing id parameter"}), 400

    where    = f"OBJECTID = {rec_id}"
    features = fetch_features(where, max_records=1)
    if not features:
        return jsonify({"error": "Record not found"}), 404

    empty_f = {k: "" for k in [
        "counties","towns","date_from","date_to","price_low","price_high",
        "land_low","land_high","street","interest","building","seller_use",
        "buyer_use","ptt_exemption","grand_list","seller_entity","seller_last",
        "seller_first","buyer_entity","buyer_last","buyer_first","span",
        "dev_prev_conv","buyer_adjoining","enrolled_current_use","foreclosed",
    ]}
    rec = feature_to_record(features[0], empty_f)
    if not rec:
        return jsonify({"error": "Could not parse record"}), 500

    # ---- Helpers ----
    def fmtd(epoch_ms):
        if not epoch_ms: return ""
        try: return _dt.fromtimestamp(int(epoch_ms)/1000, tz=_tz.utc).strftime("%m/%d/%Y")
        except: return ""

    def fmtm(v):
        if v is None or v == "": return ""
        try: return f"{int(float(v)):,}"
        except: return str(v)

    def fmtm2(v):
        """Format for J-block PDF fields: plain number, 2 decimal places, NO commas.
        PDF form fields with numeric format masks reject comma-formatted strings (show 1.#R)."""
        if v is None or v == "": return ""
        try: return f"{float(v):.2f}"
        except: return str(v)

    def boolval(v):
        if v is None: return None
        s = str(v).strip().upper()
        if s in ("TRUE","YES","1"): return "Yes"
        if s in ("FALSE","NO","0"): return "No"
        return None

    def time_held(acq_ms, close_ms):
        if not acq_ms or not close_ms: return "",""
        try:
            d1 = _dt.fromtimestamp(int(acq_ms)/1000, tz=_tz.utc)
            d2 = _dt.fromtimestamp(int(close_ms)/1000, tz=_tz.utc)
            mt = (d2.year-d1.year)*12 + (d2.month-d1.month)
            if d2.day < d1.day: mt -= 1
            return str(mt//12), str(mt%12)
        except: return "",""

    def fmt_span(raw):
        if not raw: return ""
        s = str(raw).replace("-","").zfill(11)
        return f"{s[0:3]}-{s[3:6]}-{s[6:11]}" if len(s)>=11 else s

    def bc_code(val):
        if val is None: return ""
        try: return str(int(float(str(val)))).zfill(2)
        except: return str(val)

    def strip_prefix(s):
        if not s: return ""
        return re.sub(r'^\d{2}\.\s*','',str(s)).strip()

    # ---- Derived values ----
    years_held, months_held = time_held(rec.get("dateSellerAcquired"), rec.get("closingDate"))

    fin_map = {"Conventional/Bank":"Bank","Owner Financing":"Owner","Other":"Other"}
    fin_val = fin_map.get(str(rec.get("financing") or ""), None)

    land_size = rec.get("landSize")
    no_land   = (not land_size or float(land_size) == 0) if land_size is not None else True

    # ---- AcroForm text fields ----
    fields = {
        # A — Seller
        "A-EntityName":      rec.get("sellerEntityName") or "",
        "A-LastName":        rec.get("sellerLastName")   or "",
        "A-FirstName":       rec.get("sellerFirstName")  or "",
        # B — Buyer
        "B-EntityName":      rec.get("buyerEntityName")  or "",
        "B-EntityName-COPY": rec.get("buyerEntityName")  or "",
        "B-LastName":        rec.get("buyerLastName")    or "",
        "B-LastName-COPY":   rec.get("buyerLastName")    or "",
        "B-FirstName":       rec.get("buyerFirstName")   or "",
        "B-FirstName-COPY":  rec.get("buyerFirstName")   or "",
        # C — Property
        "C-PropertyAddress":      rec.get("propertyLocationStreet") or "",
        "C-PropertyAddress-COPY": rec.get("propertyLocationStreet") or "",
        "C-PropertyCity":         rec.get("propertyLocationCity")   or "",
        "C-PropertyCity-COPY":    rec.get("propertyLocationCity")   or "",
        "C-LandSize":             str(land_size) if land_size else "",
        "C-SPAN":                 fmt_span(rec.get("span")),
        # D — Dates
        "D-YearAcquired":      fmtd(rec.get("dateSellerAcquired")),
        "D-DateOfClosing":     fmtd(rec.get("closingDate")),
        "D-DateOfClosing-COPY":fmtd(rec.get("closingDate")),
        "D-TimeHeld-Years":    years_held,
        "D-TimeHeld-Months":   months_held,
        # E — Exemptions
        "E1": str(rec.get("propertyTaxExemption") or ""),
        "E2": str(rec.get("familyMember")          or ""),
        "E3": str(rec.get("LGTExemption")          or ""),
        # F — Transfer
        "F1":  bc_code(rec.get("sellerAcquire")),
        "F1a": strip_prefix(rec.get("sellerAcquireOther")),
        "F2":  bc_code(rec.get("interestPropertyType")),
        "F2a": str(rec.get("interestUndivPercent") or ""),
        "F2b": strip_prefix(rec.get("interestPropertyTypeOther")),
        "F3-1": bc_code(rec.get("buildingConstruction1")),
        "F3-2": bc_code(rec.get("buildingConstruction2")),
        "F3-3": bc_code(rec.get("buildingConstruction3")),
        "F3a":  str(rec.get("buildingConstructionUnits05") or ""),
        "F3b":  str(rec.get("buildingConstructionDwellingUnits06") or ""),
        "F3c":  str(rec.get("buildingConstruction20") or ""),
        "F5c":  "",
        # H — Use
        "H1":  bc_code(rec.get("sellerUseOfProperty")),
        "H1a": str(rec.get("sellerUseOfPropertyExplain") or ""),
        "H2":  bc_code(rec.get("buyerUseOfProperty")),
        "H2a": str(rec.get("buyerUseOfPropertyExplain") or ""),
        # I — Withholding: leave blank
        "I2": "", "I2a": "",
        # J — Tax: J8/J9/J10/J14/J15 drawn via ReportLab overlay (Comb+JS fields
        # reject formatted strings from update_page_form_field_values — show 1.#R).
        # Leave them out of this dict entirely; overlay handles them on page 3.
        "J1":"","J2":"","J3":"","J4":"","J5":"","J6":"","J7":"",
        "J11":"","J12":"","J13":"",
        "WebKey": "",
    }

    # ---- Load and fill PDF ----
    pdf_path = os.path.join(os.path.dirname(__file__), "PTT-172-2023.pdf")
    if not os.path.exists(pdf_path):
        return jsonify({"error": "PTT-172-2023.pdf not found in app directory"}), 500

    reader = PdfReader(pdf_path)
    writer = PdfWriter()
    writer.append(reader)

    # Only update text (/Tx) fields — never pass button fields to this method
    # or it writes the value as visible text instead of checking the box.
    reader2 = PdfReader(pdf_path)
    btn_field_names = set()
    for name, f in (reader2.get_fields() or {}).items():
        if f.get('/FT') == '/Btn':
            btn_field_names.add(name)
    text_fields = {k: v for k, v in fields.items() if k not in btn_field_names}

    for pg in writer.pages:
        writer.update_page_form_field_values(pg, text_fields)

    # ---- Radio / checkbox buttons ----
    bool_fields = {
        "F4": boolval(rec.get("tenantPurchase")),
        "F6": None,
        "G1": boolval(rec.get("enrolledCurrentUse")),
        "G2": boolval(rec.get("currentUseEnrollmentContinue")),
        "H3": boolval(rec.get("rentedBefore")),
        "H4": boolval(rec.get("rentedAfter")),
        "H5": boolval(rec.get("developmentPrevConv")),
        "H6": boolval(rec.get("buyerAdjoiningProperty")),
        "H7": None,
        "I1": None,
    }

    def set_radio(wr, field_name, chosen):
        """Set a radio/checkbox field by updating both widget /AS and parent /V."""
        if not chosen: return
        for page in wr.pages:
            if "/Annots" not in page: continue
            for aref in page["/Annots"]:
                annot = aref.get_object()
                parent_ref = annot.get("/Parent")
                if not parent_ref: continue
                p = parent_ref.get_object()
                if p.get("/T") == field_name:
                    ap = annot.get("/AP", {})
                    if hasattr(ap, 'get'):
                        n = ap.get("/N", {})
                        if hasattr(n, 'keys') and f"/{chosen}" in n:
                            # Update widget appearance state
                            aref.get_object().update(
                                {NameObject("/AS"): NameObject(f"/{chosen}")}
                            )
                            # Also update parent field value so it saves correctly
                            p.update(
                                {NameObject("/V"): NameObject(f"/{chosen}")}
                            )

    for fn, val in bool_fields.items():
        set_radio(writer, fn, val)
    if fin_val:
        set_radio(writer, "F5", fin_val)

    # ---- C-DidNotInvolveLand checkbox ----
    if no_land:
        for page in writer.pages:
            if "/Annots" not in page: continue
            for aref in page["/Annots"]:
                annot = aref.get_object()
                if annot.get("/T") == "C-DidNotInvolveLand":
                    annot.update({NameObject("/AS"): NameObject("/Yes"),
                                  NameObject("/V"):  NameObject("/Yes")})

    # ---- Page 3 J-block overlay (J8/J9/J10/J14/J15 — bypass Comb field JS) ----
    # These fields have /Ff=Comb + /AA JavaScript that conflict with pypdf string writes.
    # Draw values directly as right-aligned text inside the field boxes instead.
    # Field box positions (from pdfplumber, converted to ReportLab bottom-origin):
    #   J8,J9,J10,J11,J12,J13: x0=272 x1=407  (right edge x=405)
    #   J14,J15:                x0=435 x1=571  (right edge x=569)
    # Y positions (RL bottom-origin = 792 - pdfplumber_bottom):
    #   J8:  RL_y=304   J9: RL_y=276   J10: RL_y=247
    #   J14: RL_y=134   J15: RL_y=98
    j_overlay_buf = _io.BytesIO()
    jc = rl_canvas.Canvas(j_overlay_buf, pagesize=letter)
    jc.setFont("Helvetica", 9)

    def draw_j(text, right_x, y):
        """Draw right-aligned text at the given right edge x and baseline y."""
        if text:
            jc.drawRightString(right_x, y, str(text))

    j8_val  = fmtm2(rec.get("ValuePaidOrTransferred"))
    j9_val  = fmtm2(rec.get("PersonalPropValuePaidOrTrans")) or "0.00"
    j10_val = fmtm2(rec.get("RealPropValuePaidOrTrans"))
    j14_val = fmtm2(rec.get("GenRateTaxDue"))
    j15_val = fmtm2(rec.get("TotalTaxDue"))

    draw_j(j8_val,  405, 306)
    draw_j(j9_val,  405, 278)
    draw_j(j10_val, 405, 249)
    draw_j(j14_val, 569, 136)
    draw_j(j15_val, 569, 100)

    jc.save()
    j_overlay_buf.seek(0)
    j_overlay_reader = _PR2(j_overlay_buf)
    writer.pages[2].merge_page(j_overlay_reader.pages[0])

    # ---- Page 4 town clerk overlay (not fillable fields — draw as text) ----
    gl_cat_map = {
        "1":"01","01":"01","2":"02","02":"02","3":"03","03":"03",
        "4":"04","04":"04","5":"05","05":"05","6":"06","06":"06",
        "7":"07","07":"07","8":"08","08":"08","9":"09","09":"09",
        "10":"10","11":"11","12":"12","13":"13","14":"14","15":"15",
    }
    gl_cat_raw = str(rec.get("TownGrandListCategory") or "")
    gl_cat = gl_cat_map.get(gl_cat_raw, gl_cat_raw)
    subdiv = str(rec.get("TownSubdivision") or "").strip().upper() in ("TRUE","YES","1")

    # Build an overlay page with ReportLab then merge onto page 4
    overlay_buf = _io.BytesIO()
    c = rl_canvas.Canvas(overlay_buf, pagesize=letter)
    W, H = letter   # 612 x 792 pt
    c.setFont("Helvetica", 9)

    # Page 4 coordinates (points from bottom-left, 612x792 page).
    # The town clerk table is in the middle of page 4 (below preparer section).
    # Positions measured from the rendered PDF image:
    #   Row 1 (Book / Page / GL Year):    y ~ 430
    #   Row 2 (City / Parcel / Date):      y ~ 398
    #   Row 3 (GL Value / Category / SPAN):y ~ 366
    #   Subdivision checkbox (c Subdivision): y ~ 305, x ~ 272
    def draw(x, y, text):
        if text:
            c.drawString(x, y, str(text))

    # Row 1 — Book Number / Page Number / Grand List Year
    draw(72,  400, rec.get("TownBookNumber")    or "")
    draw(275, 400, rec.get("TownPageNumber")    or "")
    draw(435, 400, rec.get("TownGrandListYear") or "")
    # Row 2 — City or Town / Parcel ID / Date of Record
    draw(72,  375, rec.get("TownCityorTown")    or "")
    draw(275, 375, rec.get("TownParcelIDNo")    or "")
    draw(435, 375, fmtd(rec.get("TownDateOfRecord")))
    # Row 3 — Grand List Value / Grand List Category / SPAN
    draw(72,  350, fmtm(rec.get("TownGrandListValue")))
    draw(275, 350, gl_cat)
    draw(435, 350, fmt_span(rec.get("TownSpan")))

    # Subdivision checkbox — the 'c Subdivision' box is the middle of 3 checkboxes.
    # Confirmed coordinates via pdfplumber: checkbox x0=261, y0=300.9 (ReportLab bottom-origin).
    # drawString at x=261, y=303 places the X cleanly inside the 14pt box.
    if subdiv:
        c.setFont("Helvetica-Bold", 11)
        c.drawString(261, 303, "X")
        c.setFont("Helvetica", 9)

    c.save()
    overlay_buf.seek(0)

    overlay_reader = _PR2(overlay_buf)
    page4 = writer.pages[3]
    page4.merge_page(overlay_reader.pages[0])

    # ---- Page 5: Full Code Reference (mirrors sidebar exactly) ----
    notes_buf = _io.BytesIO()
    nc = rl_canvas.Canvas(notes_buf, pagesize=letter)
    NW, NH = letter

    def _cd(table, val):
        """Look up code description from VT_CODES."""
        if val is None or str(val).strip() == "": return ""
        try: key = str(int(float(str(val))))
        except: key = str(val)
        return VT_CODES.get(table, {}).get(key, "")

    def _cs(val):
        """Format code as zero-padded 2-digit string."""
        if val is None or str(val).strip() == "": return ""
        try: return str(int(float(str(val)))).zfill(2)
        except: return str(val)

    def _yn(val):
        """Format boolean as Yes/No."""
        if val is True or str(val).upper() in ("TRUE","YES","1"): return "Yes"
        return "No"

    # ---- Header ----
    nc.setFillColorRGB(0.122, 0.306, 0.475)  # #1F4E79
    nc.rect(36, NH-68, NW-72, 32, fill=1, stroke=0)
    nc.setFillColorRGB(1,1,1)
    nc.setFont("Helvetica-Bold", 13)
    nc.drawString(44, NH-52, "PTT-172 Code Reference")
    nc.setFont("Helvetica", 8)
    prop_addr = ((rec.get("propertyLocationStreet") or "") + ", " +
                 (rec.get("propertyLocationCity") or "")).strip(", ")
    nc.drawString(44, NH-63, f"{prop_addr}  |  Closing: {fmtd(rec.get('closingDate'))}")
    nc.setFillColorRGB(0,0,0)

    # ---- Drawing helpers ----
    y = NH - 90
    MARGIN_L, MARGIN_R = 36, NW - 36
    COL_LINE, COL_CODE, COL_DESC = 44, 90, 120

    def section_hdr(title):
        nonlocal y
        if y < NH - 90:  y -= 4
        nc.setFillColorRGB(0.88, 0.93, 0.97)
        nc.rect(MARGIN_L, y-2, NW-72, 13, fill=1, stroke=0)
        nc.setFont("Helvetica-Bold", 8)
        nc.setFillColorRGB(0.122, 0.306, 0.475)
        nc.drawString(COL_LINE, y+2, title)
        nc.setFillColorRGB(0,0,0)
        y -= 16

    def draw_row(line, code_str, desc, plain=False):
        """Draw one data row. plain=True skips the code box."""
        nonlocal y
        check_y()
        nc.setFont("Helvetica-Bold", 7.5)
        nc.drawString(COL_LINE, y, line)
        nc.setFont("Helvetica", 7.5)
        if plain:
            nc.drawString(COL_CODE, y, str(desc))
        else:
            if code_str:
                # Blue chip box
                nc.setFillColorRGB(0.91, 0.94, 0.99)
                nc.setStrokeColorRGB(0.70, 0.78, 0.96)
                cw = nc.stringWidth(code_str, "Helvetica-Bold", 7.5) + 8
                nc.roundRect(COL_CODE, y-1, cw, 10, 2, fill=1, stroke=1)
                nc.setFillColorRGB(0.1, 0.23, 0.36)
                nc.setFont("Helvetica-Bold", 7.5)
                nc.drawString(COL_CODE+4, y, code_str)
                nc.setFillColorRGB(0,0,0)
                nc.setFont("Helvetica", 7.5)
                # Wrap description
                _wrap_text(nc, desc, COL_DESC + cw + 4, y, MARGIN_R - COL_DESC - cw - 4)
            else:
                _wrap_text(nc, desc, COL_CODE, y, MARGIN_R - COL_CODE)
        y -= 13

    def _wrap_text(canvas, text, x, start_y, max_w):
        nonlocal y
        if not text: return
        words = text.split()
        lines_out, cur = [], ""
        for w in words:
            test = (cur + " " + w).strip()
            if canvas.stringWidth(test, "Helvetica", 7.5) <= max_w:
                cur = test
            else:
                if cur: lines_out.append(cur)
                cur = w
        if cur: lines_out.append(cur)
        for i, ln in enumerate(lines_out):
            canvas.drawString(x, start_y - i*10, ln)
        if len(lines_out) > 1:
            y -= (len(lines_out)-1) * 10

    def check_y():
        nonlocal y
        if y < 55:
            nc.setFont("Helvetica", 7)
            nc.setFillColorRGB(0.5,0.5,0.5)
            nc.drawString(44, 36, "PTT-172 Code Reference (continued) — VT Property Sales")
            nc.setFillColorRGB(0,0,0)
            nc.showPage()
            y = NH - 40

    # ---- Section E ----
    section_hdr("Section E — Exemptions")
    draw_row("E1", _cs(rec.get("propertyTaxExemption")),
             _cd("ptt_exemptions", rec.get("propertyTaxExemption")) or "No exemption")
    fm = rec.get("familyMember")
    if fm and str(fm).strip() not in ("","0","0.0"):
        draw_row("E2", _cs(fm),
                 _cd("family_member_codes", fm) or rec.get("familyMemberDesc",""))
    draw_row("E3", _cs(rec.get("LGTExemption")),
             _cd("land_gains_exemptions", rec.get("LGTExemption")) or "None or no land")

    # ---- Section F ----
    section_hdr("Section F — Transfer Information")
    draw_row("F1", _cs(rec.get("sellerAcquire")),
             _cd("how_acquired_codes", rec.get("sellerAcquire")) or rec.get("sellerAcquireDesc",""))
    draw_row("F2", _cs(rec.get("interestPropertyType")),
             _cd("interest_types", rec.get("interestPropertyType")) or rec.get("interestUndivPercentDesc",""))
    for bk, bdk, lbl in [
        ("buildingConstruction1","buildingConstruction1Desc","F3"),
        ("buildingConstruction2","buildingConstruction2Desc","F3-2"),
        ("buildingConstruction3","buildingConstruction3Desc","F3-3"),
    ]:
        bv = rec.get(bk)
        if bv:
            draw_row(lbl, _cs(bv),
                     _cd("building_types", bv) or rec.get(bdk,""))
    draw_row("F4", "", f"Tenant prior to transfer: {_yn(rec.get('tenantPurchase'))}", plain=True)
    draw_row("F5", "", f"Financing: {rec.get('financing') or '—'}", plain=True)

    # ---- Section G ----
    section_hdr("Section G — Current Use")
    draw_row("G1", "", f"Enrolled in Current Use: {_yn(rec.get('enrolledCurrentUse'))}", plain=True)
    draw_row("G2", "", f"Continue enrollment: {_yn(rec.get('currentUseEnrollmentContinue'))}", plain=True)

    # ---- Section H ----
    section_hdr("Section H — Use of Property")
    draw_row("H1", _cs(rec.get("sellerUseOfProperty")),
             _cd("use_of_property", rec.get("sellerUseOfProperty")) or rec.get("sellerUseOfPropertyDesc",""))
    draw_row("H2", _cs(rec.get("buyerUseOfProperty")),
             _cd("use_of_property", rec.get("buyerUseOfProperty")) or rec.get("buyerUseOfPropertyDesc",""))
    draw_row("H3", "", f"Rented before transfer: {_yn(rec.get('rentedBefore'))}", plain=True)
    draw_row("H4", "", f"Rented after transfer: {_yn(rec.get('rentedAfter'))}", plain=True)
    draw_row("H5", "", f"Development rights conveyed separately: {_yn(rec.get('developmentPrevConv'))}", plain=True)
    draw_row("H6", "", f"Buyer holds adjoining property: {_yn(rec.get('buyerAdjoiningProperty'))}", plain=True)

    # ---- Section J ----
    section_hdr("Section J — Tax Calculation")
    draw_row("J8",  "", f"Value Paid or Transferred:  ${fmtm2(rec.get('ValuePaidOrTransferred'))}", plain=True)
    draw_row("J9",  "", f"Personal Property Value:     ${fmtm2(rec.get('PersonalPropValuePaidOrTrans')) or '0.00'}", plain=True)
    draw_row("J10", "", f"Real Property Value:         ${fmtm2(rec.get('RealPropValuePaidOrTrans'))}", plain=True)
    draw_row("J14", "", f"General Rate Tax Due:        ${fmtm2(rec.get('GenRateTaxDue'))}", plain=True)
    draw_row("J15", "", f"Total Tax Due:               ${fmtm2(rec.get('TotalTaxDue'))}", plain=True)

    # ---- Footer ----
    nc.setFont("Helvetica", 7)
    nc.setFillColorRGB(0.5,0.5,0.5)
    nc.drawString(44, 36, "PTT-172 Code Reference — VT Property Sales | Source: Form PTT-172 Instructions Rev. 11/25")
    nc.setFillColorRGB(0,0,0)

    nc.save()
    notes_buf.seek(0)
    notes_reader = _PR2(notes_buf)
    for pg in notes_reader.pages:
        writer.add_page(pg)

    # ---- Produce final PDF bytes ----
    buf = _io.BytesIO()
    writer.write(buf)
    buf.seek(0)
    pdf_bytes = buf.read()

    addr  = (rec.get("propertyLocationStreet") or "").replace(" ","_")[:30]
    close = fmtd(rec.get("closingDate")).replace("/","-")
    fname = f"PTT-172_{addr}_{close}.pdf"

    # Default: open inline in browser. If ?download=1 force download.
    # Use filename* (RFC 5987) and plain filename for maximum browser compatibility.
    if download:
        disposition = f"attachment; filename={fname}; filename*=UTF-8''{fname}"
    else:
        disposition = f"inline; filename={fname}"
    return Response(pdf_bytes, mimetype="application/pdf",
                    headers={
                        "Content-Disposition": disposition,
                        "Content-Type": "application/pdf",
                    })


@app.route("/contact", methods=["POST"])
@login_required
def contact():
    """Handle contact form submission — sends email via Resend HTTP API."""
    RESEND_API_KEY  = os.environ.get("RESEND_API_KEY", "")
    NOTIFY_EMAIL    = os.environ.get("CONTACT_NOTIFY_EMAIL", "vtrealestatesales@gmail.com")
    FROM_EMAIL      = os.environ.get("CONTACT_FROM_EMAIL", "info@vtpropertysales.com")

    data        = request.get_json() or {}
    user_email  = data.get("email", "").strip()
    subject_cat = data.get("category", "General").strip()
    message     = data.get("message", "").strip()

    if not message:
        return jsonify({"error": "Message is required."}), 400

    body = f"""From: {user_email or 'Unknown'}
Category: {subject_cat}

{message}

---
Sent via VT Property Sales contact form"""

    payload = {
        "from":     FROM_EMAIL,
        "to":       [NOTIFY_EMAIL],
        "reply_to": user_email if user_email else NOTIFY_EMAIL,
        "subject":  f"[VT Property Sales] {subject_cat}",
        "text":     body,
    }

    try:
        resp = requests.post(
            "https://api.resend.com/emails",
            headers={
                "Authorization": f"Bearer {RESEND_API_KEY}",
                "Content-Type":  "application/json",
            },
            json=payload,
            timeout=10,
        )
        if resp.status_code in (200, 201):
            return jsonify({"ok": True})
        else:
            app.logger.error(f"Resend error {resp.status_code}: {resp.text}")
            return jsonify({"error": "Failed to send message. Please try again."}), 500
    except Exception as e:
        app.logger.error(f"Contact form error: {e}")
        return jsonify({"error": "Failed to send message. Please try again."}), 500


# ── WMS Proxy ─────────────────────────────────────────────────────────────────
# Forwards WMS tile requests server-side to bypass browser CORS restrictions.
# Usage: /proxy/wms?url=<encoded_wms_url>&<wms_params>
WMS_ALLOWED = [
    'anrmaps.vermont.gov',
    'maps.vcgi.vermont.gov',
    'hazards.fema.gov',
    'services1.arcgis.com',
]

@app.route("/debug/where")
def debug_where():
    """Temporary debug endpoint to show the WHERE clause for a given filter set."""
    from flask import request as req
    filters = parse_filters(req.args)
    where = build_where_clause(filters)
    has_town = bool(filters.get('towns','').strip())
    return jsonify({'where': where, 'has_town_filter': has_town, 'filters': {k:v for k,v in filters.items() if v}})


@app.route("/proxy/wms")
def proxy_wms():
    from urllib.parse import urlparse, urlencode
    target_url = request.args.get('url', '')
    if not target_url:
        return '', 400
    parsed = urlparse(target_url)
    if not any(allowed in parsed.netloc for allowed in WMS_ALLOWED):
        return 'Forbidden', 403
    # Forward all query params except 'url'
    params = {k: v for k, v in request.args.items() if k != 'url'}
    try:
        resp = requests.get(target_url, params=params, timeout=15,
                           headers={'User-Agent': 'VTPropertyTool/1.0'})
        return resp.content, resp.status_code, {
            'Content-Type': resp.headers.get('Content-Type', 'image/png'),
            'Access-Control-Allow-Origin': '*',
            'Cache-Control': 'public, max-age=3600'
        }
    except Exception as e:
        app.logger.error(f'WMS proxy error: {e}')
        return '', 502


if __name__ == "__main__":
    app.run(debug=True)
